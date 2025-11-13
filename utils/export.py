"""
数据导出工具
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from config import Config


def prepare_work_experiences(work_experience):
    """获取最多两段工作经历并格式化"""
    experiences = work_experience or []
    top_two = experiences[:2]
    while len(top_two) < 2:
        top_two.append({})

    prepared = []
    for exp in top_two:
        start_year = exp.get('start_year') if exp else None
        end_year = exp.get('end_year') if exp else None
        if start_year and end_year:
            time_range = f"{start_year}-{end_year}"
        elif start_year and not end_year:
            time_range = f"{start_year}-至今"
        elif not start_year and end_year:
            time_range = f"至{end_year}"
        else:
            time_range = ''
        prepared.append({
            'company': (exp or {}).get('company') or '',
            'position': (exp or {}).get('position') or '',
            'start_year': start_year or '',
            'end_year': end_year or '',
            'time': time_range
        })
    return prepared


def export_resume_to_excel(resume):
    """导出单个简历到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "简历信息"
    
    # 标题样式
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 写入标题
    ws['A1'] = '简历信息'
    ws['A1'].font = title_font
    ws.merge_cells('A1:B1')
    
    # 基本信息
    row = 3
    ws[f'A{row}'] = '基本信息'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    top_exps = prepare_work_experiences(resume.work_experience)
    exp1, exp2 = top_exps

    # 计算工龄
    current_year = datetime.now().year
    earliest_year = resume.earliest_work_year
    # 如果earliest_work_year为空，从工作经历中计算
    if not earliest_year and resume.work_experience:
        work_years_list = [exp.get('start_year') for exp in resume.work_experience if exp.get('start_year')]
        if work_years_list:
            earliest_year = min(work_years_list)
    work_years = None
    if earliest_year:
        work_years = current_year - earliest_year
        if work_years < 0:
            work_years = None
    
    data = [
        ['应聘岗位', resume.applied_position or ''],
        ['姓名', resume.name or ''],
        ['性别', resume.gender or ''],
        ['出生年份', resume.birth_year or ''],
        ['年龄', resume.age or ''],
        ['工龄', work_years if work_years is not None else ''],
        ['工作经历一 - 公司', exp1['company']],
        ['工作经历一 - 岗位', exp1['position']],
        ['工作经历一 - 时间', exp1['time']],
        ['工作经历二 - 公司', exp2['company']],
        ['工作经历二 - 岗位', exp2['position']],
        ['工作经历二 - 时间', exp2['time']],
        ['手机号', resume.phone or ''],
        ['邮箱', resume.email or ''],
        ['最高学历', resume.highest_education or ''],
        ['学校（原始）', resume.school_original or ''],
        ['学校（标准化）', resume.school or ''],
        ['学校匹配状态', resume.school_match_status or ''],
        ['学校置信度', resume.school_confidence or ''],
        ['专业（原始）', resume.major_original or ''],
        ['专业（标准化）', resume.major or ''],
        ['专业匹配状态', resume.major_match_status or ''],
        ['专业置信度', resume.major_confidence or ''],
    ]
    
    for item in data:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # 工作经历
    if resume.work_experience:
        row += 1
        ws[f'A{row}'] = '工作经历'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # 表头
        headers = ['公司名称（原始）', '公司名称（标准化）', '职位', '开始年份', '结束年份', '匹配状态', '置信度']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # 数据
        for exp in resume.work_experience:
            ws.cell(row=row, column=1, value=exp.get('company', ''))
            ws.cell(row=row, column=2, value=exp.get('company_standardized', ''))
            ws.cell(row=row, column=3, value=exp.get('position', ''))
            ws.cell(row=row, column=4, value=exp.get('start_year', ''))
            ws.cell(row=row, column=5, value=exp.get('end_year', ''))
            ws.cell(row=row, column=6, value=exp.get('company_match_status', ''))
            ws.cell(row=row, column=7, value=exp.get('company_confidence', ''))
            row += 1
    
    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    
    # 保存文件
    filename = f'resume_{resume.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    file_path = os.path.join(Config.EXPORT_FOLDER, filename)
    wb.save(file_path)
    
    return file_path

def export_resumes_to_excel(resumes):
    """批量导出简历到Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "简历列表"

    def format_work_experience(experiences):
        if not experiences:
            return ''
        lines = []
        for exp in experiences:
            company = exp.get('company_standardized') or exp.get('company') or ''
            position = exp.get('position') or ''
            start_year = exp.get('start_year')
            end_year = exp.get('end_year')
            time_range = ''
            if start_year and end_year:
                time_range = f"{start_year}-{end_year}"
            elif start_year and not end_year:
                time_range = f"{start_year}-至今"
            elif not start_year and end_year:
                time_range = f"至{end_year}"
            parts = [company.strip(), position.strip(), time_range]
            line = ' | '.join([part for part in parts if part])
            if line:
                lines.append(line)
        return '\n'.join(lines)
    
    # 表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 表头
    headers = [
        'ID', '应聘岗位', '姓名', '性别', '年龄', '工龄', '手机号',
        '工作经历一（公司）', '工作经历一（岗位）', '工作经历一（时间）',
        '工作经历二（公司）', '工作经历二（岗位）', '工作经历二（时间）',
        '邮箱', '最高学历',
        '学校（原始）', '学校（标准化）', '学校状态',
        '专业（原始）', '专业（标准化）', '专业状态',
        '上传时间', '解析时间'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 数据
    current_year = datetime.now().year
    for row_idx, resume in enumerate(resumes, 2):
        ws.cell(row=row_idx, column=1, value=resume.id)
        ws.cell(row=row_idx, column=2, value=resume.applied_position or '')
        ws.cell(row=row_idx, column=3, value=resume.name or '')
        ws.cell(row=row_idx, column=4, value=resume.gender or '')
        ws.cell(row=row_idx, column=5, value=resume.age or '')
        # 计算工龄：今年-最早工作年份
        earliest_year = resume.earliest_work_year
        # 如果earliest_work_year为空，从工作经历中计算
        if not earliest_year and resume.work_experience:
            work_years_list = [exp.get('start_year') for exp in resume.work_experience if exp.get('start_year')]
            if work_years_list:
                earliest_year = min(work_years_list)
        work_years = None
        if earliest_year:
            work_years = current_year - earliest_year
            if work_years < 0:
                work_years = None
        ws.cell(row=row_idx, column=6, value=work_years if work_years is not None else '')

        exp1, exp2 = prepare_work_experiences(resume.work_experience)
        ws.cell(row=row_idx, column=7, value=exp1['company'])
        ws.cell(row=row_idx, column=8, value=exp1['position'])
        ws.cell(row=row_idx, column=9, value=exp1['time'])
        ws.cell(row=row_idx, column=10, value=exp2['company'])
        ws.cell(row=row_idx, column=11, value=exp2['position'])
        ws.cell(row=row_idx, column=12, value=exp2['time'])

        ws.cell(row=row_idx, column=13, value=resume.email or '')
        ws.cell(row=row_idx, column=14, value=resume.highest_education or '')
        ws.cell(row=row_idx, column=15, value=resume.school_original or '')
        ws.cell(row=row_idx, column=16, value=resume.school or '')
        ws.cell(row=row_idx, column=17, value=resume.school_match_status or '')
        ws.cell(row=row_idx, column=18, value=resume.major_original or '')
        ws.cell(row=row_idx, column=19, value=resume.major or '')
        ws.cell(row=row_idx, column=20, value=resume.major_match_status or '')
        ws.cell(row=row_idx, column=21, value=resume.upload_time.strftime('%Y-%m-%d %H:%M:%S') if resume.upload_time else '')
        ws.cell(row=row_idx, column=22, value=resume.parse_time.strftime('%Y-%m-%d %H:%M:%S') if resume.parse_time else '')
 
    widths = {
        1: 8, 2: 18, 3: 15, 4: 8, 5: 10, 6: 10, 7: 15,
        8: 22, 9: 18, 10: 18,
        11: 22, 12: 18, 13: 18,
        14: 22, 15: 12, 16: 18, 17: 18,
        18: 12, 19: 18, 20: 18,
        21: 12, 22: 19, 23: 19
    }
    for col in range(1, len(headers) + 1):
        width = widths.get(col, 15)
        ws.column_dimensions[chr(64 + col)].width = width
    
    # 保存文件
    filename = f'resumes_batch_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    file_path = os.path.join(Config.EXPORT_FOLDER, filename)
    wb.save(file_path)
    
    return file_path

