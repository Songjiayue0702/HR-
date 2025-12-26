"""
智能简历数据库系统 - 主应用
"""
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import io
import json
from werkzeug.utils import secure_filename
import os
import secrets
from datetime import datetime
from functools import wraps
from config import Config
from models import get_db_session, Resume, Position, Interview, User, GlobalAIConfig
from database_manager import get_database_manager
from utils.file_parser import extract_text
from utils.info_extractor import InfoExtractor
from utils.ai_extractor import AIExtractor
from utils.duplicate_checker import check_duplicate
from utils.export import export_resumes_to_excel, export_interviews_to_excel
from utils.export_pdf import export_resume_analysis_to_pdf, export_interview_round_analysis_to_pdf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
import threading
from sqlalchemy import and_
from sqlalchemy.orm import make_transient
import traceback
import sys


center_wrap_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _stylize_cell(cell, border, font=None, align=None, fill=None):
    cell.border = border
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    else:
        cell.alignment = center_wrap_alignment
    if fill:
        cell.fill = fill


def _fill_row(ws, row, values, border, align):
    ws.row_dimensions[row].height = 25
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = align
        cell.border = border


app = Flask(__name__, 
            template_folder='templates',
            static_folder='static')
app.config.from_object(Config)

# 注册中文字体
pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))

# 数据库初始化状态
db_initialized = False

def ensure_database_initialized():
    """确保数据库已初始化（延迟初始化）"""
    global db_initialized
    if db_initialized:
        return True
    
    try:
        from models import init_database, migrate_database
        init_database()
        migrate_database()
        db_initialized = True
        return True
    except Exception as e:
        print(f"警告: 数据库初始化失败: {e}")
        import traceback
        traceback.print_exc()
        return False

# 在应用启动时初始化数据库（立即执行，不等待第一个请求）
try:
    ensure_database_initialized()
    print("✓ 数据库初始化完成")
except Exception as e:
    print(f"⚠️  数据库初始化警告: {e}")
    print("应用将继续启动，数据库将在首次使用时初始化")

# OCR功能已移除，所有文档通过AI API处理

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


# ============================================================================
# AI配置管理辅助函数
# ============================================================================

def get_effective_ai_config():
    """
    获取当前有效的AI配置（优先级：用户session > 全局配置 > 环境变量）
    
    Returns:
        dict: AI配置字典，包含：
            - ai_enabled: bool
            - ai_api_key: str
            - ai_api_base: str
            - ai_model: str
    """
    config = {
        'ai_enabled': True,
        'ai_api_key': '',
        'ai_api_base': '',
        'ai_model': 'gpt-3.5-turbo'
    }
    
    # 优先级1: 用户session配置（临时配置）
    if 'ai_config' in session and isinstance(session['ai_config'], dict):
        user_config = session['ai_config']
        if user_config.get('ai_enabled') is not None:
            config['ai_enabled'] = bool(user_config.get('ai_enabled'))
        if user_config.get('ai_api_key'):
            config['ai_api_key'] = user_config.get('ai_api_key', '')
        if user_config.get('ai_api_base'):
            config['ai_api_base'] = user_config.get('ai_api_base', '')
        if user_config.get('ai_model'):
            config['ai_model'] = user_config.get('ai_model', 'gpt-3.5-turbo')
        return config
    
    # 优先级2: 全局配置（管理员设置，存储在数据库）
    db = get_db_session()
    try:
        global_config = db.query(GlobalAIConfig).first()
        if global_config:
            from utils.encryption import decrypt_value
            config['ai_enabled'] = bool(global_config.ai_enabled)
            config['ai_api_key'] = decrypt_value(global_config.ai_api_key) if global_config.ai_api_key else ''
            config['ai_api_base'] = global_config.ai_api_base or ''
            config['ai_model'] = global_config.ai_model or 'gpt-3.5-turbo'
            return config
    finally:
        db.close()
    
    # 优先级3: 环境变量（默认配置）
    config['ai_enabled'] = Config.AI_ENABLED
    config['ai_api_key'] = Config.AI_API_KEY
    config['ai_api_base'] = Config.AI_API_BASE
    config['ai_model'] = Config.AI_MODEL
    
    return config

def create_ai_extractor(ai_config=None):
    """
    创建AI提取器实例
    
    Args:
        ai_config: 可选的AI配置字典，如果为None则使用get_effective_ai_config()
    
    Returns:
        AIExtractor实例或None（如果AI未启用）
    """
    if ai_config is None:
        ai_config = get_effective_ai_config()
    
    # 检查AI是否启用
    if not ai_config.get('ai_enabled', True):
        return None
    
    # 检查API密钥
    api_key = ai_config.get('ai_api_key', '')
    if not api_key:
        return None
    
    try:
        return AIExtractor(
            api_key=api_key,
            api_base=ai_config.get('ai_api_base', ''),
            model=ai_config.get('ai_model', 'gpt-3.5-turbo')
        )
    except Exception as e:
        print(f"创建AI提取器失败: {e}")
        return None


@app.route('/')
def index():
    """
    首页 - 智能重定向
    - 未登录用户：重定向到登录页面
    - 已登录用户：重定向到应用主页面
    """
    # 检查是否已登录
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    else:
        return redirect(url_for('app_index'))

@app.route('/system-status')
def system_status():
    """系统状态页面 - 展示架构和数据库状态"""
    try:
        db_manager = get_database_manager()
        db_status = db_manager.get_status()
        test_result = db_manager.test_connection()
        
        # 检查环境变量
        env_info = {
            'PORT': os.environ.get('PORT', '未设置'),
            'HOST': os.environ.get('HOST', '未设置'),
            'CF_ACCOUNT_ID': '已设置' if os.environ.get('CF_ACCOUNT_ID') else '未设置',
            'CF_D1_DATABASE_ID': '已设置' if os.environ.get('CF_D1_DATABASE_ID') else '未设置',
            'CF_API_TOKEN': '已设置' if os.environ.get('CF_API_TOKEN') else '未设置',
            'DATABASE_PATH': os.environ.get('DATABASE_PATH', '使用默认值'),
        }
        
        return render_template('status.html', 
                             db_status=db_status,
                             test_result=test_result,
                             env_info=env_info)
    except Exception as e:
        return f"""
        <html>
        <head><title>系统状态</title></head>
        <body>
            <h1>系统状态</h1>
            <p>错误: {str(e)}</p>
            <p><a href="/health">健康检查</a></p>
            <p><a href="/api/status">API 状态</a></p>
        </body>
        </html>
        """, 500

@app.route('/health', methods=['GET'])
def health_check():
    """健康检查端点，供 Railway 等平台监控"""
    try:
        # 确保数据库已初始化
        ensure_database_initialized()
        
        # 使用 DatabaseManager 测试连接
        db_manager = get_database_manager()
        test_result = db_manager.test_connection()
        
        if test_result['success']:
            return jsonify({
                'status': 'healthy',
                'database': 'connected',
                'db_type': test_result.get('db_type', 'unknown'),
                'timestamp': datetime.now().isoformat()
            }), 200
        else:
            return jsonify({
                'status': 'unhealthy',
                'database': 'disconnected',
                'error': test_result.get('error', 'Unknown error'),
                'timestamp': datetime.now().isoformat()
            }), 503
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'database': 'disconnected',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 503

# 权限装饰器（需要在路由之前定义）
def login_required(f):
    """要求登录的装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'success': False, 'message': '请先登录', 'require_login': True}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """要求管理员权限的装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'success': False, 'message': '请先登录', 'require_login': True}), 401
        db = get_db_session()
        try:
            user = db.query(User).filter_by(id=session['user_id']).first()
            if not user or user.role != 'admin':
                return jsonify({'success': False, 'message': '需要管理员权限'}), 403
        finally:
            db.close()
        return f(*args, **kwargs)
    return decorated_function

@app.route('/api/database/init', methods=['POST'])
@admin_required
def init_database_api():
    """手动初始化数据库（管理员权限）"""
    try:
        success = ensure_database_initialized()
        if success:
            return jsonify({
                'success': True,
                'message': '数据库初始化成功'
            })
        else:
            return jsonify({
                'success': False,
                'message': '数据库初始化失败，请查看日志'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'数据库初始化失败: {str(e)}'
        }), 500

@app.route('/api/status', methods=['GET'])
def api_status():
    """API 状态信息"""
    try:
        db_manager = get_database_manager()
        db_status = db_manager.get_status()
        test_result = db_manager.test_connection()
        
        # 环境信息
        env_info = {
            'PORT': os.environ.get('PORT', '未设置'),
            'HOST': os.environ.get('HOST', '未设置'),
            'CF_ACCOUNT_ID': '已设置' if os.environ.get('CF_ACCOUNT_ID') else '未设置',
            'CF_D1_DATABASE_ID': '已设置' if os.environ.get('CF_D1_DATABASE_ID') else '未设置',
            'CF_API_TOKEN': '已设置' if os.environ.get('CF_API_TOKEN') else '未设置',
            'CF_R2_ACCOUNT_ID': '已设置' if os.environ.get('CF_R2_ACCOUNT_ID') else '未设置',
            'CF_R2_ACCESS_KEY_ID': '已设置' if os.environ.get('CF_R2_ACCESS_KEY_ID') else '未设置',
            'CF_R2_SECRET_ACCESS_KEY': '已设置' if os.environ.get('CF_R2_SECRET_ACCESS_KEY') else '未设置',
            'DATABASE_PATH': os.environ.get('DATABASE_PATH', '使用默认值'),
        }
        
        return jsonify({
            'success': True,
            'app': {
                'name': '智能简历数据库系统',
                'version': '1.0.0',
                'status': 'running'
            },
            'database': {
                'status': db_status,
                'test': test_result
            },
            'environment': env_info,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

# ============================================================================
# AI配置管理API - 管理员全局配置
# ============================================================================

@app.route('/api/admin/ai-config', methods=['GET'])
@admin_required
def get_global_ai_config():
    """获取全局AI配置（管理员权限）"""
    try:
        db = get_db_session()
        try:
            global_config = db.query(GlobalAIConfig).first()
            if global_config:
                return jsonify({
                    'success': True,
                    'data': global_config.to_dict(include_key=True)
                })
            else:
                # 返回默认配置
                return jsonify({
                    'success': True,
                    'data': {
                        'id': None,
                        'ai_enabled': Config.AI_ENABLED,
                        'ai_api_key': '',
                        'ai_api_key_set': False,
                        'ai_api_base': Config.AI_API_BASE,
                        'ai_model': Config.AI_MODEL,
                        'created_by': None,
                        'updated_by': None,
                        'created_at': None,
                        'updated_at': None
                    }
                })
        finally:
            db.close()
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取配置失败: {str(e)}'
        }), 500

@app.route('/api/admin/ai-config', methods=['POST'])
@admin_required
def set_global_ai_config():
    """设置全局AI配置（管理员权限）"""
    try:
        data = request.json or {}
        
        # 验证必填字段
        ai_enabled = data.get('ai_enabled', True)
        ai_model = data.get('ai_model', 'gpt-3.5-turbo')
        
        # 获取当前用户
        current_user = get_current_user()
        username = current_user.username if current_user else 'admin'
        
        db = get_db_session()
        try:
            # 查找或创建全局配置（单例模式）
            global_config = db.query(GlobalAIConfig).first()
            
            if not global_config:
                global_config = GlobalAIConfig()
                global_config.created_by = username
                db.add(global_config)
            
            # 更新配置
            global_config.ai_enabled = 1 if ai_enabled else 0
            global_config.ai_model = ai_model
            global_config.ai_api_base = data.get('ai_api_base', '')
            global_config.updated_by = username
            
            # 处理API密钥（加密存储）
            if 'ai_api_key' in data:
                api_key = data.get('ai_api_key', '').strip()
                if api_key:
                    from utils.encryption import encrypt_value
                    global_config.ai_api_key = encrypt_value(api_key)
                elif api_key == '':
                    # 如果传入空字符串，清除密钥
                    global_config.ai_api_key = None
            
            db.commit()
            
            return jsonify({
                'success': True,
                'message': '全局AI配置已保存',
                'data': global_config.to_dict(include_key=False)
            })
        finally:
            db.close()
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'保存配置失败: {str(e)}'
        }), 500

@app.route('/api/admin/ai-config/test', methods=['POST'])
@admin_required
def test_global_ai_config():
    """测试全局AI配置连接（管理员权限）"""
    try:
        data = request.json or {}
        
        # 构建测试配置
        test_config = {
            'ai_enabled': data.get('ai_enabled', True),
            'ai_api_key': data.get('ai_api_key', ''),
            'ai_api_base': data.get('ai_api_base', ''),
            'ai_model': data.get('ai_model', 'gpt-3.5-turbo')
        }
        
        # 创建AI提取器并测试
        ai_extractor = create_ai_extractor(test_config)
        if not ai_extractor:
            return jsonify({
                'success': False,
                'message': 'AI未启用或API密钥为空'
            }), 400
        
        # 执行简单测试
        try:
            test_result = ai_extractor.extract_info("测试文本：姓名张三，年龄25岁")
            if test_result:
                return jsonify({
                    'success': True,
                    'message': 'AI配置测试成功',
                    'test_result': test_result
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'AI配置测试失败：未返回结果'
                }), 400
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'AI配置测试失败: {str(e)}'
            }), 400
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'测试失败: {str(e)}'
        }), 500

# ============================================================================
# AI配置管理API - 用户个人配置
# ============================================================================

@app.route('/api/user/ai-config', methods=['GET'])
@login_required
def get_user_ai_config():
    """获取用户个人AI配置（登录用户）"""
    try:
        user_config = session.get('ai_config', {})
        return jsonify({
            'success': True,
            'data': user_config if user_config else {
                'ai_enabled': None,
                'ai_api_key': '',
                'ai_api_base': '',
                'ai_model': ''
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取配置失败: {str(e)}'
        }), 500

@app.route('/api/user/ai-config', methods=['POST'])
@login_required
def set_user_ai_config():
    """设置用户个人AI配置（登录用户，存储在session）"""
    try:
        data = request.json or {}
        
        # 更新session中的配置
        user_config = {
            'ai_enabled': data.get('ai_enabled'),
            'ai_api_key': data.get('ai_api_key', '').strip(),
            'ai_api_base': data.get('ai_api_base', '').strip(),
            'ai_model': data.get('ai_model', 'gpt-3.5-turbo')
        }
        
        # 只保存非空值
        session['ai_config'] = {k: v for k, v in user_config.items() if v is not None and v != ''}
        
        return jsonify({
            'success': True,
            'message': '个人AI配置已保存',
            'data': session.get('ai_config', {})
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'保存配置失败: {str(e)}'
        }), 500

@app.route('/api/user/ai-config', methods=['DELETE'])
@login_required
def clear_user_ai_config():
    """清除用户个人AI配置（登录用户）"""
    try:
        if 'ai_config' in session:
            del session['ai_config']
        
        return jsonify({
            'success': True,
            'message': '个人AI配置已清除，将使用全局配置或环境变量'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'清除配置失败: {str(e)}'
        }), 500

@app.route('/api/user/ai-config/test', methods=['POST'])
@login_required
def test_user_ai_config():
    """测试用户个人AI配置连接（登录用户）"""
    try:
        data = request.json or {}
        
        # 构建测试配置（优先使用请求中的配置，否则使用session中的配置）
        if data:
            test_config = {
                'ai_enabled': data.get('ai_enabled', True),
                'ai_api_key': data.get('ai_api_key', ''),
                'ai_api_base': data.get('ai_api_base', ''),
                'ai_model': data.get('ai_model', 'gpt-3.5-turbo')
            }
        else:
            # 使用当前有效的配置
            test_config = get_effective_ai_config()
        
        # 创建AI提取器并测试
        ai_extractor = create_ai_extractor(test_config)
        if not ai_extractor:
            return jsonify({
                'success': False,
                'message': 'AI未启用或API密钥为空'
            }), 400
        
        # 执行简单测试
        try:
            test_result = ai_extractor.extract_info("测试文本：姓名张三，年龄25岁")
            if test_result:
                return jsonify({
                    'success': True,
                    'message': 'AI配置测试成功',
                    'test_result': test_result
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'AI配置测试失败：未返回结果'
                }), 400
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'AI配置测试失败: {str(e)}'
            }), 400
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'测试失败: {str(e)}'
        }), 500

@app.route('/api/database/status', methods=['GET'])
def database_status():
    """获取数据库状态（兼容旧接口）"""
    try:
        db_manager = get_database_manager()
        db_status = db_manager.get_status()
        
        # 检查表是否存在
        tables_status = {}
        tables_to_check = ['resumes', 'positions', 'interviews', 'users']
        
        if db_status.get('tables'):
            for table_name in tables_to_check:
                tables_status[table_name] = 'exists' if table_name in db_status['tables'] else 'missing'
        else:
            # 如果无法获取表列表，尝试直接查询
            db = get_db_session()
            if db:
                from sqlalchemy import text
                for table_name in tables_to_check:
                    try:
                        db.execute(text(f'SELECT 1 FROM {table_name} LIMIT 1'))
                        tables_status[table_name] = 'exists'
                    except Exception:
                        tables_status[table_name] = 'missing'
                db.close()
        
        return jsonify({
            'success': True,
            'initialized': db_status.get('initialized', False),
            'db_type': db_status.get('db_type', 'unknown'),
            'tables': tables_status,
            'all_tables_exist': all(status == 'exists' for status in tables_status.values())
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/init-db', methods=['GET', 'POST'])
def init_database_route():
    """初始化数据库表"""
    try:
        from models import init_database, migrate_database
        
        # 初始化数据库
        init_database()
        migrate_database()
        
        # 获取数据库状态
        db_manager = get_database_manager()
        db_status = db_manager.get_status()
        
        return jsonify({
            'success': True,
            'message': '数据库初始化成功',
            'status': db_status,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'数据库初始化失败: {str(e)}',
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/test-d1', methods=['GET'])
def test_d1_connection():
    """测试 D1 数据库连接"""
    try:
        db_manager = get_database_manager()
        test_result = db_manager.test_connection()
        
        return jsonify({
            'success': test_result['success'],
            'db_type': test_result.get('db_type', 'unknown'),
            'message': test_result.get('message', ''),
            'error': test_result.get('error'),
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/env-check', methods=['GET'])
def env_check():
    """检查环境变量配置"""
    env_vars = {
        'PORT': os.environ.get('PORT'),
        'HOST': os.environ.get('HOST'),
        'SECRET_KEY': '已设置' if os.environ.get('SECRET_KEY') else '未设置',
        'DEBUG': os.environ.get('DEBUG', 'False'),
        # Cloudflare D1
        'CF_ACCOUNT_ID': os.environ.get('CF_ACCOUNT_ID'),
        'CF_D1_DATABASE_ID': os.environ.get('CF_D1_DATABASE_ID'),
        'CF_API_TOKEN': '已设置' if os.environ.get('CF_API_TOKEN') else '未设置',
        # Cloudflare R2
        'CF_R2_ACCOUNT_ID': os.environ.get('CF_R2_ACCOUNT_ID'),
        'CF_R2_ACCESS_KEY_ID': os.environ.get('CF_R2_ACCESS_KEY_ID'),
        'CF_R2_SECRET_ACCESS_KEY': '已设置' if os.environ.get('CF_R2_SECRET_ACCESS_KEY') else '未设置',
        'CF_R2_BUCKET_NAME': os.environ.get('CF_R2_BUCKET_NAME'),
        # 数据库
        'DATABASE_PATH': os.environ.get('DATABASE_PATH'),
        # AI 配置
        'AI_ENABLED': os.environ.get('AI_ENABLED', 'true'),
        'AI_API_KEY': '已设置' if os.environ.get('OPENAI_API_KEY') or os.environ.get('AI_API_KEY') else '未设置',
        'AI_MODEL': os.environ.get('AI_MODEL', 'gpt-3.5-turbo'),
    }
    
    # 检查关键配置
    checks = {
        'railway_configured': bool(os.environ.get('PORT')),
        'd1_configured': bool(os.environ.get('CF_D1_DATABASE_ID') and os.environ.get('CF_ACCOUNT_ID')),
        'r2_configured': bool(os.environ.get('CF_R2_ACCOUNT_ID') and os.environ.get('CF_R2_ACCESS_KEY_ID')),
        'database_path_set': bool(os.environ.get('DATABASE_PATH')),
    }
    
    return jsonify({
        'success': True,
        'environment_variables': env_vars,
        'configuration_checks': checks,
        'recommendations': _get_env_recommendations(checks),
        'timestamp': datetime.now().isoformat()
    })

def _get_env_recommendations(checks: dict) -> list:
    """获取环境配置建议"""
    recommendations = []
    
    if not checks['railway_configured']:
        recommendations.append('建议设置 PORT 环境变量（Railway 会自动设置）')
    
    if not checks['d1_configured']:
        recommendations.append('如需使用 Cloudflare D1，请设置 CF_ACCOUNT_ID 和 CF_D1_DATABASE_ID')
    
    if not checks['r2_configured']:
        recommendations.append('如需使用 Cloudflare R2，请设置 CF_R2_ACCOUNT_ID 和 CF_R2_ACCESS_KEY_ID')
    
    if not os.environ.get('SECRET_KEY'):
        recommendations.append('建议设置 SECRET_KEY 环境变量以提高安全性')
    
    return recommendations

@app.route('/api/education-levels', methods=['GET'])
def get_education_levels():
    return jsonify({
        'success': True,
        'data': Config.EDUCATION_LEVELS
    })

def process_resume_async(resume_id, file_path):
    """异步处理简历解析"""
    db = get_db_session()
    try:
        resume = db.query(Resume).filter_by(id=resume_id).first()
        if not resume:
            return
        
        resume.parse_status = 'processing'
        db.commit()
        
        # 提取文本
        raw_text = extract_text(file_path)
        if not raw_text:
            raise Exception("无法从文件中提取文本，文件可能已损坏或格式不支持")
        
        # 检测文件类型
        file_ext = os.path.splitext(file_path)[1].lower()
        is_word_file = file_ext in ['.doc', '.docx']
        
        # 异步任务使用全局配置（不依赖session）
        # 优先级：全局配置 > 环境变量
        db_config = get_db_session()
        try:
            global_config = db_config.query(GlobalAIConfig).first()
            if global_config:
                from utils.encryption import decrypt_value
                ai_enabled = bool(global_config.ai_enabled)
                ai_api_key = decrypt_value(global_config.ai_api_key) if global_config.ai_api_key else ''
                ai_api_base = global_config.ai_api_base or ''
                ai_model = global_config.ai_model or 'gpt-3.5-turbo'
            else:
                # 使用环境变量
                ai_enabled = Config.AI_ENABLED
                ai_api_key = Config.AI_API_KEY
                ai_api_base = Config.AI_API_BASE
                ai_model = Config.AI_MODEL
        finally:
            db_config.close()
        
        # 如果链接了AI API，优先使用AI优化文本提取
        text = raw_text
        ai_extractor = None
        if ai_enabled and ai_api_key:
            try:
                ai_extractor = create_ai_extractor({
                    'ai_enabled': ai_enabled,
                    'ai_api_key': ai_api_key,
                    'ai_api_base': ai_api_base,
                    'ai_model': ai_model
                })
                # 使用AI优化文本提取
                optimized_text = ai_extractor.optimize_text_extraction(raw_text)
                if optimized_text:
                    text = optimized_text
                    print(f"AI文本优化成功（模型: {ai_model}），文本长度: {len(text)} 字符")
                else:
                    print(f"AI文本优化失败（模型: {ai_model}），使用原始文本")
            except Exception as e:
                print(f"AI文本优化失败（模型: {ai_model}），使用原始文本: {e}")
        
        # 保存原始文本（优先保存优化后的文本，如果没有优化则保存原始文本）
        # 注意：raw_text字段存储的是用于信息提取的文本（可能是优化后的）
        # 如果需要保存真正的原始文本，可以添加新字段
        resume.raw_text = text
        
        # 如果使用了AI优化，原始文本和优化后的文本都保存
        # 但为了信息提取准确性，使用优化后的文本进行提取
        
        # 规则提取
        extractor = InfoExtractor()
        
        # AI辅助信息提取（如果启用）
        ai_result = None
        if ai_enabled and ai_api_key and ai_extractor:
            try:
                ai_result = ai_extractor.extract_with_ai(text, is_word_file=is_word_file)
                if ai_result:
                    print(f"AI辅助信息提取成功（模型: {ai_model}，Word格式: {is_word_file}），提取到 {len([k for k, v in ai_result.items() if v])} 个字段")
            except Exception as e:
                print(f"AI辅助信息提取失败（模型: {ai_model}），继续使用规则提取: {e}")
        
        # 融合规则提取和AI提取的结果
        info = extractor.extract_all(text, ai_result=ai_result)
        
        # 更新基本信息
        resume.name = info.get('name')
        resume.gender = info.get('gender')
        resume.birth_year = info.get('birth_year')
        # 如果从简历中提取到了年龄，保存到age_from_resume
        extracted_age = info.get('age')
        if extracted_age:
            resume.age_from_resume = extracted_age
            resume.age = extracted_age
        else:
            # 如果没有提取到年龄，但有出生年份，计算年龄
            if resume.birth_year:
                resume.age = datetime.now().year - resume.birth_year
        resume.phone = info.get('phone')
        resume.email = info.get('email')
        resume.highest_education = info.get('highest_education')
        resume.raw_text = text
        resume.error_message = None
        
        # 处理工作经历（统一使用AI API智能识别，无需外部验证）
        work_experiences = info.get('work_experience', [])
        resume.work_experience = work_experiences
        
        # 处理学校信息（仅保留原文提取）
        school_original = info.get('school')
        if school_original:
            resume.school = school_original
            resume.school_original = school_original
        
        # 处理专业信息（仅保留原文提取）
        major_original = info.get('major')
        if major_original:
            resume.major = major_original
            resume.major_original = major_original
        
        # 计算并保存最早工作年份
        if work_experiences:
            work_years = [exp.get('start_year') for exp in work_experiences if exp.get('start_year')]
            if work_years:
                resume.earliest_work_year = min(work_years)
        
        # 查重检测
        existing_resumes = db.query(Resume).filter(
            Resume.parse_status == 'success',
            Resume.id != resume.id
        ).all()
        duplicate_id, similarity = check_duplicate(resume, existing_resumes)
        
        if similarity >= 80.0:
            resume.duplicate_status = '重复简历'
            resume.duplicate_similarity = similarity
            resume.duplicate_resume_id = duplicate_id
        else:
            resume.duplicate_status = None
            resume.duplicate_similarity = similarity if similarity > 0 else None
            resume.duplicate_resume_id = None
        
        resume.parse_status = 'success'
        resume.parse_time = datetime.now()
        db.commit()
        
    except Exception as e:
        resume.parse_status = 'failed'
        resume.error_message = str(e)
        db.commit()
        print(f"处理简历失败: {e}")
    finally:
        db.close()

def get_current_user():
    """获取当前登录用户"""
    if 'user_id' not in session:
        return None
    db = get_db_session()
    try:
        return db.query(User).filter_by(id=session['user_id']).first()
    finally:
        db.close()

@app.route('/app')
def app_index():
    """应用首页（需要登录）"""
    # 检查是否已登录
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    return render_template('index.html')

@app.route('/login', methods=['GET'])
def login_page():
    """登录页面"""
    # 如果已登录，重定向到应用主页面
    if 'user_id' in session:
        return redirect(url_for('app_index'))
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
def login():
    """登录接口"""
    data = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    
    if not username or not password:
        return jsonify({'success': False, 'message': '用户名和密码不能为空'}), 400
    
    db = get_db_session()
    try:
        user = db.query(User).filter_by(username=username).first()
        if not user or not user.check_password(password):
            return jsonify({'success': False, 'message': '用户名或密码错误'}), 401
        
        if user.is_active != 1:
            return jsonify({'success': False, 'message': '账户已被禁用'}), 403
        
        # 登录成功，设置session
        session['user_id'] = user.id
        session['username'] = user.username
        session['role'] = user.role
        session['real_name'] = user.real_name
        
        return jsonify({
            'success': True,
            'message': '登录成功',
            'user': user.to_dict()
        })
    finally:
        db.close()

@app.route('/api/logout', methods=['POST'])
def logout():
    """登出接口"""
    session.clear()
    return jsonify({'success': True, 'message': '已退出登录'})

@app.route('/api/current-user', methods=['GET'])
def get_current_user_info():
    """获取当前用户信息"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    db = get_db_session()
    try:
        user = db.query(User).filter_by(id=session['user_id']).first()
        if not user:
            session.clear()
            return jsonify({'success': False, 'message': '用户不存在'}), 401
        
        return jsonify({
            'success': True,
            'user': user.to_dict()
        })
    finally:
        db.close()


@app.route('/api/statistics', methods=['GET'])
@login_required
def get_statistics():
    """
    获取数据统计（需要登录）

    查询参数：
    - start_date: 开始日期（YYYY-MM-DD）
    - end_date: 结束日期（YYYY-MM-DD）
    - position: 岗位名称（可选，为空则统计全部岗位）
    - scope: 统计范围（personal-个人, department-部门, group-小组, all-整体），根据角色自动限制

    统计内容：
    - resume_count: 简历数（按简历上传时间统计）
    - interview_count: 到面数（按一面时间统计，有一面时间视为到面）
    - pass_count: 通过数（按状态为"面试通过/已发offer/已入职"等统计）
    - offer_count: offer数（按offer发放时间统计）
    - onboard_count: 入职数（按入职时间统计）
    """
    # 获取当前用户
    current_user = get_current_user()
    if not current_user:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    
    db_session = get_db_session()
    try:
        start_date_str = request.args.get('start_date', '').strip()
        end_date_str = request.args.get('end_date', '').strip()
        position = request.args.get('position', '').strip()
        scope = request.args.get('scope', '').strip()  # 统计范围
        
        # 根据角色确定可用的统计范围
        available_scopes = []
        if current_user.role == 'admin':
            # 管理员：可选个人/小组/整体
            available_scopes = ['personal', 'group', 'all']
            if not scope or scope not in available_scopes:
                scope = 'all'  # 默认整体
        elif current_user.role == 'manager':
            # 主管：可选个人/部门
            available_scopes = ['personal', 'department']
            if not scope or scope not in available_scopes:
                scope = 'department'  # 默认部门
        else:  # employee
            # 员工：只看个人
            available_scopes = ['personal']
            scope = 'personal'

        # 解析日期
        start_dt = None
        end_dt = None
        if start_date_str:
            try:
                start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'success': False, 'message': '开始日期格式错误，应为YYYY-MM-DD'}), 400
        if end_date_str:
            try:
                # 结束日期取当天 23:59:59
                end_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
                end_dt = end_dt.replace(hour=23, minute=59, second=59)
            except ValueError:
                return jsonify({'success': False, 'message': '结束日期格式错误，应为YYYY-MM-DD'}), 400

        # 1) 简历数（按上传时间）
        resume_query = db_session.query(Resume)
        if start_dt is not None:
            resume_query = resume_query.filter(Resume.upload_time >= start_dt)
        if end_dt is not None:
            resume_query = resume_query.filter(Resume.upload_time <= end_dt)
        if position:
            resume_query = resume_query.filter(Resume.applied_position == position)
        resume_count = resume_query.count()

        # 2) 到面数（有一面时间视为到面，round1_time 非空）
        interview_query = db_session.query(Interview)
        if position:
            interview_query = interview_query.filter(Interview.applied_position == position)

        # 一面时间是字符串 YYYY-MM-DD，所以用字符串范围过滤
        if start_date_str:
            interview_query = interview_query.filter(
                Interview.round1_time.isnot(None),
                Interview.round1_time >= start_date_str
            )
        if end_date_str:
            interview_query = interview_query.filter(
                Interview.round1_time <= end_date_str
            )
        interview_count = interview_query.count()

        # 3) 通过数（状态为"面试通过/已发offer/已入职"）
        pass_statuses = ['面试通过', '已发offer', '已入职']
        pass_query = db_session.query(Interview).filter(Interview.status.in_(pass_statuses))
        if position:
            pass_query = pass_query.filter(Interview.applied_position == position)
        # 通过数暂不按时间字段细分，统一按 create_time 范围过滤
        if start_dt is not None:
            pass_query = pass_query.filter(Interview.create_time >= start_dt)
        if end_dt is not None:
            pass_query = pass_query.filter(Interview.create_time <= end_dt)
        pass_count = pass_query.count()

        # 4) Offer 数（按 offer_date 字符串日期）
        offer_query = db_session.query(Interview).filter(
            Interview.offer_issued == 1,
            Interview.offer_date.isnot(None),
            Interview.offer_date != ''
        )
        if position:
            offer_query = offer_query.filter(Interview.applied_position == position)
        if start_date_str:
            offer_query = offer_query.filter(Interview.offer_date >= start_date_str)
        if end_date_str:
            offer_query = offer_query.filter(Interview.offer_date <= end_date_str)
        offer_count = offer_query.count()

        # 5) 入职数（按 onboard_date 字符串日期）
        onboard_query = db_session.query(Interview).filter(
            Interview.onboard == 1,
            Interview.onboard_date.isnot(None),
            Interview.onboard_date != ''
        )
        if position:
            onboard_query = onboard_query.filter(Interview.applied_position == position)
        if start_date_str:
            onboard_query = onboard_query.filter(Interview.onboard_date >= start_date_str)
        if end_date_str:
            onboard_query = onboard_query.filter(Interview.onboard_date <= end_date_str)
        onboard_count = onboard_query.count()

        # 如果指定了岗位，返回单个岗位的数据
        if position:
            return jsonify({
                'success': True,
                'data': {
                    'position': position,
                    'resume_count': resume_count,
                    'interview_count': interview_count,
                    'pass_count': pass_count,
                    'offer_count': offer_count,
                    'onboard_count': onboard_count,
                }
            })
        
        # 返回统计范围信息
        result_data = {
            'scope': scope,
            'available_scopes': available_scopes,
            'user_role': current_user.role
        }
        
        # 如果指定了岗位，返回单个岗位的数据
        if position:
            result_data.update({
                'position': position,
                'resume_count': resume_count,
                'interview_count': interview_count,
                'pass_count': pass_count,
                'offer_count': offer_count,
                'onboard_count': onboard_count,
            })
            return jsonify({
                'success': True,
                'data': result_data
            })
        
        # 如果没有指定岗位，按岗位分组返回数据
        from sqlalchemy import func
        stats_by_position = []
        
        # 获取所有有数据的岗位（从简历和面试记录中）
        all_positions = set()
        # 从简历中获取岗位
        resume_positions = db_session.query(Resume.applied_position).filter(
            Resume.applied_position.isnot(None),
            Resume.applied_position != ''
        )
        if start_dt is not None:
            resume_positions = resume_positions.filter(Resume.upload_time >= start_dt)
        if end_dt is not None:
            resume_positions = resume_positions.filter(Resume.upload_time <= end_dt)
        for pos in resume_positions.distinct():
            if pos[0]:
                all_positions.add(pos[0])
        
        # 从面试记录中获取岗位
        interview_positions = db_session.query(Interview.applied_position).filter(
            Interview.applied_position.isnot(None),
            Interview.applied_position != ''
        )
        if start_date_str:
            interview_positions = interview_positions.filter(
                Interview.round1_time.isnot(None),
                Interview.round1_time >= start_date_str
            )
        if end_date_str:
            interview_positions = interview_positions.filter(
                Interview.round1_time <= end_date_str
            )
        for pos in interview_positions.distinct():
            if pos[0]:
                all_positions.add(pos[0])
        
        # 为每个岗位计算统计数据
        for pos_name in sorted(all_positions):
            # 简历数
            pos_resume_query = db_session.query(Resume).filter(Resume.applied_position == pos_name)
            if start_dt is not None:
                pos_resume_query = pos_resume_query.filter(Resume.upload_time >= start_dt)
            if end_dt is not None:
                pos_resume_query = pos_resume_query.filter(Resume.upload_time <= end_dt)
            pos_resume_count = pos_resume_query.count()
            
            # 到面数
            pos_interview_query = db_session.query(Interview).filter(Interview.applied_position == pos_name)
            if start_date_str:
                pos_interview_query = pos_interview_query.filter(
                    Interview.round1_time.isnot(None),
                    Interview.round1_time >= start_date_str
                )
            if end_date_str:
                pos_interview_query = pos_interview_query.filter(Interview.round1_time <= end_date_str)
            pos_interview_count = pos_interview_query.count()
            
            # 通过数
            pos_pass_query = db_session.query(Interview).filter(
                Interview.applied_position == pos_name,
                Interview.status.in_(pass_statuses)
            )
            if start_dt is not None:
                pos_pass_query = pos_pass_query.filter(Interview.create_time >= start_dt)
            if end_dt is not None:
                pos_pass_query = pos_pass_query.filter(Interview.create_time <= end_dt)
            pos_pass_count = pos_pass_query.count()
            
            # Offer数
            pos_offer_query = db_session.query(Interview).filter(
                Interview.applied_position == pos_name,
                Interview.offer_issued == 1,
                Interview.offer_date.isnot(None),
                Interview.offer_date != ''
            )
            if start_date_str:
                pos_offer_query = pos_offer_query.filter(Interview.offer_date >= start_date_str)
            if end_date_str:
                pos_offer_query = pos_offer_query.filter(Interview.offer_date <= end_date_str)
            pos_offer_count = pos_offer_query.count()
            
            # 入职数
            pos_onboard_query = db_session.query(Interview).filter(
                Interview.applied_position == pos_name,
                Interview.onboard == 1,
                Interview.onboard_date.isnot(None),
                Interview.onboard_date != ''
            )
            if start_date_str:
                pos_onboard_query = pos_onboard_query.filter(Interview.onboard_date >= start_date_str)
            if end_date_str:
                pos_onboard_query = pos_onboard_query.filter(Interview.onboard_date <= end_date_str)
            pos_onboard_count = pos_onboard_query.count()
            
            stats_by_position.append({
                'position': pos_name,
                'resume_count': pos_resume_count,
                'interview_count': pos_interview_count,
                'pass_count': pos_pass_count,
                'offer_count': pos_offer_count,
                'onboard_count': pos_onboard_count,
            })
        
        # 如果没有岗位数据，返回空数组
        return jsonify({
            'success': True,
            'data': {
                'by_position': stats_by_position,
                'total': {
                    'resume_count': resume_count,
                    'interview_count': interview_count,
                    'pass_count': pass_count,
                    'offer_count': offer_count,
                    'onboard_count': onboard_count,
                }
            }
        })
    finally:
        db_session.close()

# 账号管理API
@app.route('/api/users', methods=['GET'])
@login_required
def list_users():
    """获取用户列表（需要登录）"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    
    db = get_db_session()
    try:
        # 管理员可以看到所有用户，主管和员工只能看到自己
        if current_user.role == 'admin':
            users = db.query(User).all()
        else:
            users = [current_user]
        
        return jsonify({
            'success': True,
            'data': [user.to_dict() for user in users]
        })
    finally:
        db.close()

@app.route('/api/users', methods=['POST'])
@admin_required
def create_user():
    """创建用户（仅管理员）"""
    data = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    role = data.get('role', 'employee').strip()
    real_name = data.get('real_name', '').strip()
    department = data.get('department', '').strip()
    group_name = data.get('group_name', '').strip()
    is_active = data.get('is_active', 1)  # 默认为激活状态
    
    if not username or not password:
        return jsonify({'success': False, 'message': '用户名和密码不能为空'}), 400
    
    if len(password) < 6:
        return jsonify({'success': False, 'message': '密码长度至少6位'}), 400
    
    if role not in ['admin', 'manager', 'employee']:
        return jsonify({'success': False, 'message': '角色无效'}), 400
    
    db = get_db_session()
    try:
        # 检查用户名是否已存在
        existing_user = db.query(User).filter_by(username=username).first()
        if existing_user:
            return jsonify({'success': False, 'message': '用户名已存在'}), 400
        
        # 创建新用户
        user = User(
            username=username,
            role=role,
            real_name=real_name,
            department=department,
            group_name=group_name,
            is_active=1 if is_active else 0
        )
        user.set_password(password)
        db.add(user)
        db.commit()
        
        return jsonify({
            'success': True,
            'message': '用户创建成功',
            'user': user.to_dict()
        })
    finally:
        db.close()

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
def update_user(user_id):
    """更新用户信息"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    
    # 非管理员只能修改自己的信息
    if current_user.role != 'admin' and current_user.id != user_id:
        return jsonify({'success': False, 'message': '无权修改其他用户信息'}), 403
    
    data = request.json or {}
    db = get_db_session()
    try:
        user = db.query(User).filter_by(id=user_id).first()
        if not user:
            return jsonify({'success': False, 'message': '用户不存在'}), 404
        
        # 更新字段
        if 'real_name' in data:
            user.real_name = data['real_name']
        if 'department' in data:
            user.department = data['department']
        if 'group_name' in data:
            user.group_name = data['group_name']
        
        # 只有管理员可以修改角色和激活状态
        if current_user.role == 'admin':
            if 'role' in data and data['role'] in ['admin', 'manager', 'employee']:
                user.role = data['role']
            if 'is_active' in data:
                user.is_active = 1 if data['is_active'] else 0
        
        # 修改密码
        # 管理员可以修改任意用户密码，普通用户只能修改自己的密码
        if 'password' in data and data['password']:
            if current_user.role == 'admin' or current_user.id == user_id:
                if len(data['password']) < 6:
                    return jsonify({'success': False, 'message': '密码长度至少6位'}), 400
                user.set_password(data['password'])
            else:
                return jsonify({'success': False, 'message': '无权修改其他用户的密码'}), 403
        
        db.commit()
        
        return jsonify({
            'success': True,
            'message': '更新成功',
            'user': user.to_dict()
        })
    finally:
        db.close()

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    """删除用户（仅管理员）"""
    current_user = get_current_user()
    if not current_user:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    
    # 不能删除自己
    if current_user.id == user_id:
        return jsonify({'success': False, 'message': '不能删除自己的账户'}), 400
    
    db = get_db_session()
    try:
        user = db.query(User).filter_by(id=user_id).first()
        if not user:
            return jsonify({'success': False, 'message': '用户不存在'}), 404
        
        db.delete(user)
        db.commit()
        
        return jsonify({
            'success': True,
            'message': '用户删除成功'
        })
    finally:
        db.close()

@app.route('/api/users/departments', methods=['GET'])
@login_required
def get_departments():
    """获取所有部门列表（用于下拉选择）"""
    db = get_db_session()
    try:
        departments = db.query(User.department).filter(
            User.department.isnot(None),
            User.department != ''
        ).distinct().all()
        return jsonify({
            'success': True,
            'data': [dept[0] for dept in departments if dept[0]]
        })
    finally:
        db.close()

@app.route('/api/users/groups', methods=['GET'])
@login_required
def get_groups():
    """获取所有小组列表（用于下拉选择）"""
    db = get_db_session()
    try:
        groups = db.query(User.group_name).filter(
            User.group_name.isnot(None),
            User.group_name != ''
        ).distinct().all()
        return jsonify({
            'success': True,
            'data': [group[0] for group in groups if group[0]]
        })
    finally:
        db.close()

@app.route('/api/upload', methods=['POST'])
@login_required
def upload_file():
    """文件上传接口（支持多文件上传，需要登录）"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有文件'}), 400
    
    # 获取所有上传的文件
    files = request.files.getlist('file')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'success': False, 'message': '未选择文件'}), 400
    
    # 获取AI配置（如果前端提供了）
    ai_config = request.form.get('ai_config')
    if ai_config:
        try:
            ai_config_data = json.loads(ai_config)
            # 更新临时配置（仅用于本次处理）
            if ai_config_data.get('ai_api_key'):
                app.config['AI_API_KEY'] = ai_config_data['ai_api_key']
            if ai_config_data.get('ai_api_base'):
                app.config['AI_API_BASE'] = ai_config_data['ai_api_base']
            if ai_config_data.get('ai_model'):
                app.config['AI_MODEL'] = ai_config_data['ai_model']
            if 'ai_enabled' in ai_config_data:
                app.config['AI_ENABLED'] = ai_config_data['ai_enabled']
        except:
            pass  # 如果解析失败，使用默认配置
    
    uploaded_count = 0
    failed_files = []
    resume_ids = []
    
    # 处理每个文件
    for file in files:
        if file.filename == '':
            continue
        
        if not allowed_file(file.filename):
            failed_files.append(file.filename)
            continue
        
        try:
            original_name = file.filename
            name_part, ext = os.path.splitext(original_name)
            ext = ext.lower()
            safe_name = secure_filename(name_part)
            if not safe_name:
                safe_name = 'resume'
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')  # 添加微秒确保唯一性
            filename = f"{timestamp}{safe_name}{ext}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # 创建数据库记录
            db = get_db_session()
            current_user = get_current_user()
            username = current_user.username if current_user else 'system'
            
            resume = Resume(
                file_name=file.filename,
                file_path=file_path,
                parse_status='pending',
                created_by=username,
                updated_by=username
            )
            db.add(resume)
            db.commit()
            resume_id = resume.id
            db.close()
            
            # 异步处理
            thread = threading.Thread(target=process_resume_async, args=(resume_id, file_path))
            thread.daemon = True
            thread.start()
            
            uploaded_count += 1
            resume_ids.append(resume_id)
        except Exception as e:
            print(f"文件 {file.filename} 上传失败: {e}")
            failed_files.append(file.filename)
    
    # 返回结果
    if uploaded_count > 0:
        message = f'成功上传 {uploaded_count} 个文件，正在解析...'
        if failed_files:
            message += f'，{len(failed_files)} 个文件上传失败'
        return jsonify({
            'success': True,
            'message': message,
            'uploaded_count': uploaded_count,
            'failed_count': len(failed_files),
            'failed_files': failed_files,
            'resume_ids': resume_ids
        })
    else:
        return jsonify({
            'success': False,
            'message': f'所有文件上传失败：{", ".join(failed_files) if failed_files else "不支持的文件格式"}'
        }), 400

@app.route('/api/resumes', methods=['GET'])
def get_resumes():
    """获取简历列表"""
    db = get_db_session()
    
    # 分页
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    # 筛选
    query = db.query(Resume)
    
    # 搜索
    search = request.args.get('search', '')
    if search:
        query = query.filter(
            (Resume.name.like(f'%{search}%')) |
            (Resume.school.like(f'%{search}%')) |
            (Resume.major.like(f'%{search}%'))
        )
    
    # 筛选
    gender = request.args.get('gender', '')
    if gender:
        query = query.filter(Resume.gender == gender)
    
    education = request.args.get('education', '')
    if education:
        query = query.filter(Resume.highest_education == education)
    
    # 排序
    sort_by = request.args.get('sort_by', 'upload_time')
    sort_order = request.args.get('sort_order', 'desc')
    if hasattr(Resume, sort_by):
        if sort_order == 'desc':
            query = query.order_by(getattr(Resume, sort_by).desc())
        else:
            query = query.order_by(getattr(Resume, sort_by).asc())
    
    # 允许显示所有状态的简历（pending/processing/success/failed）
    # 用户可以通过状态筛选来查看特定状态的简历
    status_filter = request.args.get('status', '')
    if status_filter:
        query = query.filter(Resume.parse_status == status_filter)
    # 如果没有指定状态筛选，默认显示所有状态的简历
    
    total = query.count()
    resumes = query.offset((page - 1) * per_page).limit(per_page).all()
    
    # 注意：查重检测在简历解析时完成，这里不需要重复检测
    # 如果需要对旧简历进行查重，可以单独运行查重脚本
    
    db.close()
    
    return jsonify({
        'success': True,
        'data': [r.to_dict() for r in resumes],
        'total': total,
        'page': page,
        'per_page': per_page
    })

@app.route('/api/resumes/<int:resume_id>', methods=['GET'])
def get_resume_detail(resume_id):
    """获取简历详情"""
    db = get_db_session()
    resume = db.query(Resume).filter_by(id=resume_id).first()
    db.close()
    
    if not resume:
        return jsonify({'success': False, 'message': '简历不存在'}), 404
    
    return jsonify({
        'success': True,
        'data': resume.to_dict()
    })

@app.route('/api/resumes/<int:resume_id>/download', methods=['GET'])
def download_resume_file(resume_id):
    """下载简历原始文件（支持预览和下载）"""
    session = get_db_session()
    try:
        resume = session.query(Resume).filter(Resume.id == resume_id).first()
        if not resume:
            return jsonify({'success': False, 'message': '简历不存在'}), 404
        
        file_path = resume.file_path
        if not file_path or not os.path.exists(file_path):
            return jsonify({'success': False, 'message': '文件不存在'}), 404
        
        # 获取原始文件名
        file_name = resume.file_name or os.path.basename(file_path)
        
        # 检查是否为预览模式（通过查询参数）
        as_attachment = request.args.get('download', 'false').lower() == 'true'
        
        # 根据文件类型设置MIME类型
        file_ext = os.path.splitext(file_path)[1].lower()
        mimetype = 'application/octet-stream'
        if file_ext == '.pdf':
            mimetype = 'application/pdf'
        elif file_ext in ['.doc', '.docx']:
            mimetype = 'application/msword' if file_ext == '.doc' else 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        
        return send_file(
            file_path,
            as_attachment=as_attachment,
            download_name=file_name if as_attachment else None,
            mimetype=mimetype
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'下载失败: {str(e)}'}), 500
    finally:
        session.close()

@app.route('/api/resumes/<int:resume_id>', methods=['PUT'])
def update_resume(resume_id):
    """更新简历信息"""
    db = get_db_session()
    resume = db.query(Resume).filter_by(id=resume_id).first()
    
    if not resume:
        db.close()
        return jsonify({'success': False, 'message': '简历不存在'}), 404
    
    data = request.json
    
    # 更新字段
    if 'name' in data:
        resume.name = data['name']
    if 'gender' in data:
        resume.gender = data['gender']
    if 'birth_year' in data:
        resume.birth_year = data['birth_year']
        # 如果简历中提取到了年龄，优先使用简历中的年龄
        if resume.age_from_resume:
            resume.age = resume.age_from_resume
        elif resume.birth_year:
            # 如果没有从简历提取到年龄，根据出生年份计算
            resume.age = datetime.now().year - resume.birth_year
    if 'earliest_work_year' in data:
        resume.earliest_work_year = data['earliest_work_year']
    if 'school' in data:
        resume.school = data['school']
    if 'school_original' in data:
        resume.school_original = data['school_original']
    if 'major' in data:
        resume.major = data['major']
    if 'major_original' in data:
        resume.major_original = data['major_original']
    if 'work_experience' in data:
        resume.work_experience = data['work_experience']
        # 工作经历更新时，如果用户没有手动设置earliest_work_year，自动从工作经历中计算
        if 'earliest_work_year' not in data and resume.work_experience:
            work_years = [exp.get('start_year') for exp in resume.work_experience if exp.get('start_year')]
            if work_years:
                resume.earliest_work_year = min(work_years)
    if 'highest_education' in data:
        resume.highest_education = data['highest_education']
    if 'phone' in data:
        resume.phone = data['phone']
    if 'email' in data:
        resume.email = data['email']
    if 'applied_position' in data:
        resume.applied_position = data['applied_position']
    if 'error_message' in data:
        resume.error_message = data['error_message']
    
    # 如果用户手动设置了earliest_work_year，使用用户设置的值
    if 'earliest_work_year' in data:
        resume.earliest_work_year = data['earliest_work_year']
    
    db.commit()
    db.close()
    
    return jsonify({'success': True, 'message': '更新成功'})


def _remove_file_if_exists(path: str) -> None:
    if path and os.path.exists(path):
        try:
            os.remove(path)
        except OSError:
            pass


@app.route('/api/resumes/<int:resume_id>', methods=['DELETE'])
@admin_required
def delete_resume(resume_id):
    """删除单个简历（仅管理员）"""
    db = get_db_session()
    resume = db.query(Resume).filter_by(id=resume_id).first()
    if not resume:
        db.close()
        return jsonify({'success': False, 'message': '简历不存在'}), 404

    # 检查是否有关联的面试流程
    interview_count = db.query(Interview).filter(Interview.resume_id == resume_id).count()
    if interview_count > 0:
        # 如果有关联的面试流程，提示用户（但不阻止删除）
        # 面试流程记录会保留，但简历信息会丢失
        # 面试流程中的冗余字段（name, applied_position）会保留
        pass

    file_path = resume.file_path
    db.delete(resume)
    db.commit()
    db.close()

    _remove_file_if_exists(file_path)

    return jsonify({'success': True, 'message': '删除成功'})


@app.route('/api/resumes/batch_delete', methods=['POST'])
@admin_required
def delete_resumes_batch():
    """批量删除简历（仅管理员）"""
    data = request.json or {}
    resume_ids = data.get('resume_ids', [])
    if not resume_ids:
        return jsonify({'success': False, 'message': '请选择要删除的简历'}), 400

    db = get_db_session()
    resumes = db.query(Resume).filter(Resume.id.in_(resume_ids)).all()

    if not resumes:
        db.close()
        return jsonify({'success': False, 'message': '没有找到匹配的简历'}), 404

    # 检查是否有关联的面试流程（仅用于提示，不阻止删除）
    interview_count = db.query(Interview).filter(Interview.resume_id.in_(resume_ids)).count()

    file_paths = [resume.file_path for resume in resumes]
    for resume in resumes:
        db.delete(resume)
    db.commit()
    db.close()

    for path in file_paths:
        _remove_file_if_exists(path)

    return jsonify({'success': True, 'message': '批量删除成功', 'deleted': len(file_paths)})

@app.route('/api/export/<int:resume_id>', methods=['GET'])
def export_single(resume_id):
    """导出单个简历"""
    from utils.export import export_resume_to_excel
    
    db = get_db_session()
    resume = db.query(Resume).filter_by(id=resume_id).first()
    db.close()
    
    if not resume:
        return jsonify({'success': False, 'message': '简历不存在'}), 404
    
    file_path = export_resume_to_excel(resume)
    return send_file(file_path, as_attachment=True, download_name=f'简历_{resume.name or resume.id}.xlsx')

@app.route('/api/export/batch', methods=['POST'])
def export_batch():
    """批量导出"""
    data = request.json
    resume_ids = data.get('resume_ids', [])
    
    db = get_db_session()
    resumes = db.query(Resume).filter(Resume.id.in_(resume_ids)).all()
    db.close()
    
    if not resumes:
        return jsonify({'success': False, 'message': '没有可导出的简历'}), 400
    
    file_path = export_resumes_to_excel(resumes)
    return send_file(file_path, as_attachment=True, download_name=f'简历批量导出_{datetime.now().strftime("%Y%m%d")}.xlsx')


@app.route('/api/resumes/<int:resume_id>/analysis-pdf', methods=['POST'])
def export_resume_analysis_pdf(resume_id):
    """导出简历分析详情为PDF（基本信息、教育信息、工作经历、简历匹配度分析）"""
    session = get_db_session()
    try:
        resume = session.query(Resume).filter(Resume.id == resume_id).first()
        if not resume:
            return jsonify({'success': False, 'message': '简历不存在'}), 404

        applied_position = (resume.applied_position or '').strip()
        analysis = None

        # 如果有应聘岗位且AI可用，则在导出前实时执行一次匹配分析，保证PDF中的匹配度内容是最新的
        try:
            if applied_position:
                # 获取岗位信息
                position = session.query(Position).filter(Position.position_name == applied_position).first()
                if position:
                    ai_enabled = app.config.get('AI_ENABLED', True)
                    ai_api_key = app.config.get('AI_API_KEY', '')
                    ai_model = app.config.get('AI_MODEL', 'gpt-3.5-turbo')
                    ai_api_base = app.config.get('AI_API_BASE', '')

                    if ai_enabled and ai_api_key:
                        ai_extractor = AIExtractor(
                            api_key=ai_api_key,
                            api_base=ai_api_base if ai_api_base else None,
                            model=ai_model
                        )

                        resume_info = f"""
姓名：{resume.name or '未知'}
性别：{resume.gender or '未知'}
年龄：{resume.age or '未知'}
学历：{resume.highest_education or '未知'}
毕业学校：{resume.school or '未知'}
专业：{resume.major or '未知'}
工龄：{resume.earliest_work_year and (datetime.now().year - resume.earliest_work_year) or '未知'}年
工作经历：{json.dumps(resume.work_experience or [], ensure_ascii=False, indent=2)}
"""

                        position_info = f"""
岗位名称：{position.position_name}
工作内容：{position.work_content or '未填写'}
任职资格：{position.job_requirements or '未填写'}
核心需求：{position.core_requirements or '未填写'}
"""

                        prompt = f"""请分析以下简历与岗位的匹配度，并给出详细的分析报告。

【简历信息】
{resume_info}

【岗位要求】
{position_info}

请从以下维度进行分析：
1. 教育背景匹配度（学历、学校、专业）
2. 工作经验匹配度（工作年限、工作内容、岗位相关性）
3. 技能匹配度（根据工作经历推断的技能）
4. 综合匹配度评分（0-100分）

请以JSON格式返回分析结果，格式如下：
{{
    "match_score": 85,
    "match_level": "高度匹配",
    "detailed_analysis": "详细的分析说明...",
    "strengths": ["优势1", "优势2"],
    "weaknesses": ["不足1", "不足2"],
    "suggestions": ["【考核重点】技术能力 - 【面试问题】请详细说明您在XX项目中的技术实现方案和遇到的挑战", "【考核重点】沟通协作 - 【面试问题】请描述一次您与跨部门团队协作解决复杂问题的经历"]
}}

其中：
- match_score: 匹配度分数（0-100）
- match_level: 匹配等级（高度匹配/中等匹配/低度匹配）
- detailed_analysis: 详细分析说明（200-500字）
- strengths: 优势匹配点列表
- weaknesses: 不足匹配点列表
- suggestions: 这是给面试官使用的面试重点考核项及对应面试问题，不是给候选人的建议！

【suggestions字段的详细要求】：
1. 这是给面试官的建议，用于指导面试官在面试中重点考核哪些方面，以及应该问什么问题
2. 绝对不要生成给候选人的改进建议（如"建议候选人如何提升"、"候选人应该做什么"等）
3. 必须严格按照以下格式生成，每个suggestion必须是：【考核重点】考核项名称 - 【面试问题】具体的面试问题
4. 根据简历与岗位的匹配情况，识别3-5个需要重点考核的维度，例如：
   - 如果简历缺乏相关经验，考核重点可以是"行业经验"或"学习能力"
   - 如果简历有相关经验但不够深入，考核重点可以是"项目深度"或"技术能力"
   - 如果岗位需要沟通能力，考核重点可以是"沟通协作"或"团队合作"
5. 为每个考核重点设计1-2个针对性的面试问题，问题要能帮助面试官深入了解候选人在该维度的真实能力
6. 面试问题应该以"请"、"请描述"、"请说明"等开头，直接面向候选人提问

【格式示例】：
正确格式：
- "【考核重点】技术能力 - 【面试问题】请详细说明您在XX项目中的技术实现方案和遇到的挑战"
- "【考核重点】沟通协作 - 【面试问题】请描述一次您与跨部门团队协作解决复杂问题的经历"

错误格式（禁止使用）：
- "建议候选人提升技术能力"（这是给候选人的建议，不是给面试官的）
- "候选人应该加强沟通能力"（这是给候选人的建议，不是给面试官的）
- "技术能力：请说明..."（缺少【考核重点】和【面试问题】标记）

请只返回JSON格式，不要包含其他文字说明。"""

                        response_text = ai_extractor._call_ai_api(prompt)

                        # 解析JSON
                        import re
                        try:
                            json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
                            if json_match:
                                analysis = json.loads(json_match.group())
                            else:
                                analysis = json.loads(response_text)
                        except json.JSONDecodeError:
                            analysis = {
                                'match_score': 50,
                                'match_level': '中等匹配',
                                'detailed_analysis': response_text[:500] if len(response_text) > 500 else response_text,
                                'strengths': [],
                                'weaknesses': [],
                                'suggestions': []
                            }
                        
                        # 验证和清理suggestions字段，确保格式正确
                        if 'suggestions' in analysis and isinstance(analysis['suggestions'], list):
                            import re
                            cleaned_suggestions = []
                            for suggestion in analysis['suggestions']:
                                if not isinstance(suggestion, str):
                                    continue
                                # 检查是否符合格式：【考核重点】xxx - 【面试问题】xxx
                                if re.match(r'【考核重点】.*?\s*[-—–]\s*【面试问题】.*', suggestion):
                                    cleaned_suggestions.append(suggestion)
                                # 如果不符合格式，尝试修复或跳过
                                # 过滤掉明显是给候选人的建议（包含"建议"、"应该"等关键词）
                                elif any(keyword in suggestion for keyword in ['建议候选人', '候选人应该', '建议您', '您应该', '可以尝试']):
                                    # 跳过给候选人的建议
                                    continue
                            analysis['suggestions'] = cleaned_suggestions

                        # 对得分做同样的"放宽"与等级划分，保持与页面一致
                        try:
                            raw_score = analysis.get('match_score')
                            if raw_score is None:
                                raw_score = 60
                            raw_score = float(raw_score)
                            new_score = int(max(50, min(100, raw_score * 0.7 + 30)))
                            analysis['match_score'] = new_score

                            if new_score >= 80:
                                level = '高度匹配'
                            elif new_score >= 60:
                                level = '中等匹配'
                            else:
                                level = '低度匹配'
                            analysis['match_level'] = level
                        except Exception:
                            pass
        except Exception as _:
            # 匹配度分析失败时，不影响PDF导出，只是不带匹配信息
            analysis = None

        file_path = export_resume_analysis_to_pdf(resume, analysis)
        # 文件名称格式：候选人姓名-简历分析报告
        candidate_name = resume.name or f"简历{resume_id}"
        download_name = f"{candidate_name}-简历分析报告.pdf"
        return send_file(file_path, as_attachment=True, download_name=download_name)
    except Exception as e:
        return jsonify({'success': False, 'message': f'导出分析报告失败: {str(e)}'}), 500
    finally:
        session.close()


@app.route('/api/interviews', methods=['GET'])
def list_interviews():
    """获取面试流程列表，可按姓名/岗位搜索"""
    session = get_db_session()
    try:
        search = (request.args.get('search') or '').strip()
        # 关联简历表以便生成身份验证码（使用LEFT JOIN处理简历不存在的情况）
        query = session.query(Interview, Resume).outerjoin(Resume, Interview.resume_id == Resume.id)
        if search:
            like = f"%{search}%"
            query = query.filter(
                (Interview.name.like(like)) |
                (Interview.applied_position.like(like))
            )
        rows = query.order_by(Interview.update_time.desc()).all()
        data = []
        for iv, res in rows:
            d = iv.to_dict()
            # 身份验证码：优先使用面试记录中存储的，如果为空则动态生成
            identity_code = iv.identity_code
            if not identity_code:
                # 如果面试记录中没有存储身份验证码，动态生成
                if res and res.name:
                    # 简历存在，使用简历信息生成身份验证码
                    phone = res.phone or ''
                    if phone and len(phone) >= 4:
                        identity_code = res.name + phone[-4:]
                    else:
                        identity_code = res.name
                    # 同时更新候选人姓名为简历中的姓名，确保一致性
                    d['name'] = res.name
                    # 更新面试记录中的身份验证码（异步更新，不阻塞返回）
                    try:
                        iv.identity_code = identity_code
                        session.commit()
                    except:
                        session.rollback()
                elif iv.name:
                    # 简历不存在，使用面试记录中的冗余姓名（无法获取手机号）
                    identity_code = iv.name
            else:
                # 如果简历存在，同步更新候选人姓名以确保一致性
                if res and res.name:
                    d['name'] = res.name
            d['identity_code'] = identity_code
            data.append(d)
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取面试列表失败: {str(e)}'}), 500
    finally:
        session.close()


@app.route('/api/interviews', methods=['POST'])
def create_interview():
    """
    创建一条面试流程记录
    请求体: { "resume_id": 123 }
    自动带入当前简历的姓名和应聘岗位
    """
    try:
        data = request.json or {}
        resume_id = data.get('resume_id')
        match_score = data.get('match_score')
        match_level = data.get('match_level')
        if not resume_id:
            return jsonify({'success': False, 'message': '缺少简历ID'}), 400

        session = get_db_session()
        try:
            resume = session.query(Resume).filter(Resume.id == resume_id).first()
            if not resume:
                return jsonify({'success': False, 'message': '简历不存在'}), 404

            # 生成身份验证码：姓名+手机号后四位
            identity_code = ''
            if resume.name:
                phone = resume.phone or ''
                if phone and len(phone) >= 4:
                    identity_code = resume.name + phone[-4:]
                else:
                    identity_code = resume.name
            
            # 如果请求中没有传递匹配度，尝试从简历记录中获取（如果岗位匹配）
            final_match_score = match_score
            final_match_level = match_level
            if not final_match_score and resume.match_score and resume.match_position:
                # 如果简历有匹配度分析结果，且岗位匹配，则使用简历中的匹配度
                if resume.match_position == (resume.applied_position or ''):
                    final_match_score = resume.match_score
                    final_match_level = resume.match_level
            
            current_user = get_current_user()
            username = current_user.username if current_user else 'system'
            
            interview = Interview(
                resume_id=resume.id,
                name=resume.name or '',
                applied_position=resume.applied_position or '',
                identity_code=identity_code,
                match_score=final_match_score,
                match_level=final_match_level,
                created_by=username,
                updated_by=username
            )
            session.add(interview)
            session.commit()

            return jsonify({'success': True, 'data': interview.to_dict()})
        finally:
            session.close()
    except Exception as e:
        return jsonify({'success': False, 'message': f'创建面试记录失败: {str(e)}'}), 500


def _calc_interview_status(interview: Interview) -> str:
    """
    根据各轮结果 + Offer/入职信息 计算面试流程状态
    基本规则：
    1. 优先级：已入职 > 已发offer > 轮次结果
    2. 轮次结果：
       - 默认：待面试
       - 一面 result 未通过：一面面试未通过
       - 一面通过，二面未填：一面面试通过
       - 二面 result 未通过：二面面试未通过
       - 二面通过，round3_enabled=0：面试通过
       - 二面通过，round3_enabled=1 且三面未填：二面面试通过
       - 三面 result 未通过：三面面试未通过
       - 三面 result 通过：面试通过
    3. Offer 与入职：
       - 如果 onboard=1 且实际入职日期、入职架构填写完整：状态为“已入职”
       - 否则，如果 offer_issued=1 且 Offer 发放日期、拟入职架构、拟入职日期填写完整：状态为“已发offer”
    """
    # 先处理入职/offer状态（最高优先级）
    if interview.onboard and interview.onboard_date and interview.onboard_department:
        return '已入职'

    if interview.offer_issued and interview.offer_date and interview.offer_department and interview.offer_onboard_plan_date:
        return '已发offer'

    # 以下为原有轮次状态计算逻辑
    # 一面
    if interview.round1_result:
        if interview.round1_result == '未通过':
            return '一面面试未通过'
        elif interview.round1_result == '通过':
            # 看二面
            if not interview.round2_result:
                return '一面面试通过'
    else:
        return '待面试'

    # 二面
    if interview.round2_result:
        if interview.round2_result == '未通过':
            return '二面面试未通过'
        elif interview.round2_result == '通过':
            if not interview.round3_enabled:
                return '面试通过'
            # 有三面
            if not interview.round3_result:
                return '二面面试通过'
    else:
        return '一面面试通过'

    # 三面
    if interview.round3_result:
        if interview.round3_result == '未通过':
            return '三面面试未通过'
        elif interview.round3_result == '通过':
            return '面试通过'

    return interview.status or '待面试'


@app.route('/api/interviews/<int:interview_id>', methods=['GET'])
def get_interview(interview_id):
    """获取单条面试记录详情"""
    session = get_db_session()
    try:
        interview = session.query(Interview).filter(Interview.id == interview_id).first()
        if not interview:
            return jsonify({'success': False, 'message': '面试记录不存在'}), 404
        
        data = interview.to_dict()
        
        # 身份验证码：优先使用面试记录中存储的，如果为空则动态生成
        identity_code = interview.identity_code
        if not identity_code:
            # 尝试获取关联的简历信息以生成身份验证码
            resume = session.query(Resume).filter(Resume.id == interview.resume_id).first()
            if resume:
                # 简历存在，使用简历信息生成身份验证码
                if resume.name:
                    phone = resume.phone or ''
                    if phone and len(phone) >= 4:
                        identity_code = resume.name + phone[-4:]
                    else:
                        identity_code = resume.name
                # 同时更新候选人姓名为简历中的姓名，确保一致性
                data['name'] = resume.name
                # 更新面试记录中的身份验证码
                try:
                    interview.identity_code = identity_code
                    session.commit()
                except:
                    session.rollback()
            else:
                # 简历不存在，使用面试记录中的冗余姓名
                identity_code = interview.name if interview.name else ''
        else:
            # 如果简历存在，同步更新候选人姓名以确保一致性
            resume = session.query(Resume).filter(Resume.id == interview.resume_id).first()
            if resume and resume.name:
                data['name'] = resume.name
        
        data['identity_code'] = identity_code
        
        return jsonify({'success': True, 'data': data})
    finally:
        session.close()


@app.route('/api/interviews/<int:interview_id>', methods=['PUT'])
def update_interview(interview_id):
    """更新面试流程详情"""
    data = request.json or {}
    session = get_db_session()
    try:
        interview = session.query(Interview).filter(Interview.id == interview_id).first()
        if not interview:
            return jsonify({'success': False, 'message': '面试记录不存在'}), 404

        # 记录更新者
        current_user = get_current_user()
        username = current_user.username if current_user else 'system'
        interview.updated_by = username

        # 更新各轮信息
        interview.round1_interviewer = data.get('round1_interviewer')
        interview.round1_time = data.get('round1_time')
        interview.round1_result = data.get('round1_result')

        interview.round2_interviewer = data.get('round2_interviewer')
        interview.round2_time = data.get('round2_time')
        interview.round2_result = data.get('round2_result')

        interview.round3_enabled = 1 if data.get('round3_enabled') else 0
        interview.round3_interviewer = data.get('round3_interviewer')
        interview.round3_time = data.get('round3_time')
        interview.round3_result = data.get('round3_result')

        # 分轮次面试评价
        interview.round1_comment = data.get('round1_comment')
        interview.round2_comment = data.get('round2_comment')
        interview.round3_comment = data.get('round3_comment')

        # AI 分析结果（只读展示，但需要持久化）
        interview.round1_ai_result = data.get('round1_ai_result')
        interview.round2_ai_result = data.get('round2_ai_result')
        interview.round3_ai_result = data.get('round3_ai_result')

        # 应聘岗位（同时同步到简历）
        applied_position = data.get('applied_position')
        if applied_position is not None:
            interview.applied_position = applied_position or None
            try:
                resume = session.query(Resume).filter(Resume.id == interview.resume_id).first()
                if resume:
                    resume.applied_position = applied_position or None
            except Exception as _:
                pass

        # 同步更新身份验证码（如果简历存在）
        try:
            resume = session.query(Resume).filter(Resume.id == interview.resume_id).first()
            if resume and resume.name:
                # 重新生成身份验证码
                phone = resume.phone or ''
                if phone and len(phone) >= 4:
                    interview.identity_code = resume.name + phone[-4:]
                else:
                    interview.identity_code = resume.name
                # 同时更新候选人姓名
                interview.name = resume.name
        except Exception as _:
            pass

        # Offer 与入职信息
        interview.offer_issued = 1 if data.get('offer_issued') else 0
        interview.offer_date = data.get('offer_date')
        interview.offer_department = data.get('offer_department')
        interview.offer_onboard_plan_date = data.get('offer_onboard_plan_date')

        interview.onboard = 1 if data.get('onboard') else 0
        interview.onboard_date = data.get('onboard_date')
        interview.onboard_department = data.get('onboard_department')

        # 自动计算状态（与身份验证码绑定）
        # 状态包括：待面试、一面面试通过/未通过、二面面试通过/未通过、三面面试未通过、面试通过、已发offer、已入职
        # 所有状态都与身份验证码绑定，确保通过身份验证码可以查询到完整的面试流程和最终状态
        interview.status = _calc_interview_status(interview)
        
        # 确保身份验证码与所有信息绑定（包括状态）
        # 身份验证码作为唯一标识，绑定所有面试流程详情（各轮面试信息、评价、结果、Offer、入职）和最终状态
        if not interview.identity_code:
            try:
                resume = session.query(Resume).filter(Resume.id == interview.resume_id).first()
                if resume and resume.name:
                    phone = resume.phone or ''
                    if phone and len(phone) >= 4:
                        interview.identity_code = resume.name + phone[-4:]
                    else:
                        interview.identity_code = resume.name
                elif interview.name:
                    interview.identity_code = interview.name
            except Exception as _:
                pass

        session.commit()
        return jsonify({'success': True, 'data': interview.to_dict()})
    except Exception as e:
        session.rollback()
        return jsonify({'success': False, 'message': f'更新面试记录失败: {str(e)}'}), 500
    finally:
        session.close()


@app.route('/api/interviews/resume-ids', methods=['GET'])
def get_interview_resume_ids():
    """返回所有已邀约面试的 resume_id 列表，用于前端标记"已邀约" """
    session = get_db_session()
    try:
        ids = session.query(Interview.resume_id).distinct().all()
        id_list = [row[0] for row in ids]
        return jsonify({'success': True, 'data': id_list})
    finally:
        session.close()


@app.route('/api/interviews/by-identity', methods=['GET'])
def get_interview_by_identity():
    """通过身份验证码查询面试流程详情"""
    identity_code = request.args.get('identity_code', '').strip()
    if not identity_code:
        return jsonify({'success': False, 'message': '缺少身份验证码'}), 400
    
    session = get_db_session()
    try:
        # 通过身份验证码查找面试记录
        interview = session.query(Interview).filter(Interview.identity_code == identity_code).first()
        if not interview:
            return jsonify({'success': False, 'message': '未找到对应的面试记录'}), 404
        
        data = interview.to_dict()
        
        # 如果简历存在，同步更新候选人姓名以确保一致性
        resume = session.query(Resume).filter(Resume.id == interview.resume_id).first()
        if resume and resume.name:
            data['name'] = resume.name
        
        return jsonify({'success': True, 'data': data})
    finally:
        session.close()


@app.route('/api/interviews/<int:interview_id>/comment-link/<int:round>', methods=['POST'])
def generate_comment_link(interview_id, round):
    """生成面试评价填写链接"""
    if round not in (1, 2, 3):
        return jsonify({'success': False, 'message': '无效的轮次'}), 400
    
    session = get_db_session()
    try:
        interview = session.query(Interview).filter(Interview.id == interview_id).first()
        if not interview:
            return jsonify({'success': False, 'message': '面试记录不存在'}), 404
        
        # 生成唯一token
        token = secrets.token_urlsafe(32)
        
        # 保存token到对应字段
        if round == 1:
            interview.round1_comment_token = token
        elif round == 2:
            interview.round2_comment_token = token
        else:
            interview.round3_comment_token = token
        
        session.commit()
        return jsonify({'success': True, 'data': {'token': token}})
    except Exception as e:
        session.rollback()
        return jsonify({'success': False, 'message': f'生成链接失败: {str(e)}'}), 500
    finally:
        session.close()


@app.route('/interview-comment', methods=['GET'])
def interview_comment_page():
    """面试评价填写页面（公开页面）"""
    token = request.args.get('token', '').strip()
    if not token:
        return render_template('error.html', message='缺少访问令牌'), 400
    
    session = get_db_session()
    try:
        # 查找对应的面试记录和轮次
        interview = None
        round_num = None
        
        interview1 = session.query(Interview).filter(Interview.round1_comment_token == token).first()
        if interview1:
            interview = interview1
            round_num = 1
        else:
            interview2 = session.query(Interview).filter(Interview.round2_comment_token == token).first()
            if interview2:
                interview = interview2
                round_num = 2
            else:
                interview3 = session.query(Interview).filter(Interview.round3_comment_token == token).first()
                if interview3:
                    interview = interview3
                    round_num = 3
        
        if not interview:
            return render_template('error.html', message='无效的访问令牌'), 404
        
        # 获取当前轮次的信息
        current_comment = ''
        current_interviewer = ''
        current_time = ''
        current_result = ''
        
        if round_num == 1:
            current_comment = interview.round1_comment or ''
            current_interviewer = interview.round1_interviewer or ''
            current_time = interview.round1_time or ''
            current_result = interview.round1_result or ''
        elif round_num == 2:
            current_comment = interview.round2_comment or ''
            current_interviewer = interview.round2_interviewer or ''
            current_time = interview.round2_time or ''
            current_result = interview.round2_result or ''
        else:
            current_comment = interview.round3_comment or ''
            current_interviewer = interview.round3_interviewer or ''
            current_time = interview.round3_time or ''
            current_result = interview.round3_result or ''
        
        return render_template('interview_comment.html', 
                             interview=interview, 
                             round=round_num,
                             current_comment=current_comment,
                             current_interviewer=current_interviewer,
                             current_time=current_time,
                             current_result=current_result,
                             token=token)
    finally:
        session.close()


@app.route('/api/interview-comment/submit', methods=['POST'])
def submit_interview_comment():
    """提交面试评价及相关信息"""
    data = request.json or {}
    token = data.get('token', '').strip()
    interviewer = data.get('interviewer', '').strip()
    time = data.get('time', '').strip()
    result = data.get('result', '').strip()
    comment = data.get('comment', '').strip()
    
    if not token:
        return jsonify({'success': False, 'message': '缺少访问令牌'}), 400
    
    session = get_db_session()
    try:
        # 查找对应的面试记录和轮次
        interview = None
        round_num = None
        
        interview1 = session.query(Interview).filter(Interview.round1_comment_token == token).first()
        if interview1:
            interview = interview1
            round_num = 1
        else:
            interview2 = session.query(Interview).filter(Interview.round2_comment_token == token).first()
            if interview2:
                interview = interview2
                round_num = 2
            else:
                interview3 = session.query(Interview).filter(Interview.round3_comment_token == token).first()
                if interview3:
                    interview = interview3
                    round_num = 3
        
        if not interview:
            return jsonify({'success': False, 'message': '无效的访问令牌'}), 404
        
        # 保存当轮次的信息
        if round_num == 1:
            interview.round1_interviewer = interviewer
            interview.round1_time = time
            interview.round1_result = result
            interview.round1_comment = comment
        elif round_num == 2:
            interview.round2_interviewer = interviewer
            interview.round2_time = time
            interview.round2_result = result
            interview.round2_comment = comment
        else:
            interview.round3_interviewer = interviewer
            interview.round3_time = time
            interview.round3_result = result
            interview.round3_comment = comment
        
        # 自动计算状态
        interview.status = _calc_interview_status(interview)
        
        session.commit()
        return jsonify({'success': True, 'message': '信息提交成功'})
    except Exception as e:
        session.rollback()
        return jsonify({'success': False, 'message': f'提交失败: {str(e)}'}), 500
    finally:
        session.close()


@app.route('/api/interviews/<int:interview_id>/upload-doc', methods=['POST'])
def upload_interview_doc(interview_id):
    """上传面试文档（录音逐字稿等），按轮次区分"""
    session = get_db_session()
    try:
        interview = session.query(Interview).filter(Interview.id == interview_id).first()
        if not interview:
            return jsonify({'success': False, 'message': '面试记录不存在'}), 404

        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '未找到上传文件'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '文件名为空'}), 400

        # 轮次：1/2/3
        round_str = request.form.get('round')
        if round_str not in ('1', '2', '3'):
            return jsonify({'success': False, 'message': '缺少或错误的轮次参数'}), 400

        # 保存到上传目录
        upload_folder = os.path.join(app.static_folder, 'interview_docs')
        os.makedirs(upload_folder, exist_ok=True)
        filename = secure_filename(f"interview_{interview_id}_round{round_str}_{file.filename}")
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)

        # 保存相对路径（供前端访问）
        rel_path = f"interview_docs/{filename}"
        if round_str == '1':
            interview.round1_doc_path = rel_path
        elif round_str == '2':
            interview.round2_doc_path = rel_path
        else:
            interview.round3_doc_path = rel_path
        session.commit()

        return jsonify({'success': True, 'data': interview.to_dict()})
    except Exception as e:
        session.rollback()
        return jsonify({'success': False, 'message': f'上传录音失败: {str(e)}'}), 500
    finally:
        session.close()


@app.route('/api/interviews/batch_delete', methods=['POST'])
@admin_required
def delete_interviews_batch():
    """批量删除面试流程（仅管理员）"""
    try:
        data = request.json or {}
        interview_ids = data.get('interview_ids', [])
        
        if not interview_ids:
            return jsonify({'success': False, 'message': '请选择要删除的面试记录'}), 400
        
        session = get_db_session()
        try:
            interviews = session.query(Interview).filter(Interview.id.in_(interview_ids)).all()
            if not interviews:
                return jsonify({'success': False, 'message': '未找到要删除的面试记录'}), 404
            
            # 删除关联的文档文件（如果有）
            for interview in interviews:
                for doc_path in [interview.round1_doc_path, interview.round2_doc_path, interview.round3_doc_path]:
                    if doc_path:
                        full_path = os.path.join(app.static_folder, doc_path)
                        if os.path.exists(full_path):
                            try:
                                os.remove(full_path)
                            except Exception as e:
                                print(f"删除文档文件失败: {e}")
            
            # 删除面试记录
            for interview in interviews:
                session.delete(interview)
            
            session.commit()
            return jsonify({
                'success': True,
                'message': f'成功删除 {len(interviews)} 条面试记录',
                'deleted': len(interviews)
            })
        finally:
            session.close()
    except Exception as e:
        return jsonify({'success': False, 'message': f'批量删除失败: {str(e)}'}), 500

@app.route('/api/interviews/export', methods=['POST'])
def export_interviews():
    """导出面试流程：选中或全部"""
    try:
        data = request.json or {}
        interview_ids = data.get('interview_ids', [])

        session = get_db_session()
        if interview_ids:
            interviews = session.query(Interview).filter(Interview.id.in_(interview_ids)).all()
        else:
            interviews = session.query(Interview).order_by(Interview.update_time.desc()).all()

        # 构造简历映射，用于导出时生成身份验证码
        resume_ids = [iv.resume_id for iv in interviews]
        resumes = session.query(Resume).filter(Resume.id.in_(resume_ids)).all() if resume_ids else []
        resume_map = {r.id: r for r in resumes}
        session.close()

        if not interviews:
            return jsonify({'success': False, 'message': '没有可导出的面试记录'}), 400

        file_path = export_interviews_to_excel(interviews, resume_map)
        return send_file(
            file_path,
            as_attachment=True,
            download_name=f'面试流程导出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500


@app.route('/api/interviews/<int:interview_id>/analyze-doc', methods=['POST'])
def analyze_interview_doc(interview_id):
    """
    使用AI对面试文档（录音逐字稿等）进行分析
    请求体: { "round": 1|2|3 }
    """
    try:
        data = request.json or {}
        round_str = str(data.get('round') or '')
        if round_str not in ('1', '2', '3'):
            return jsonify({'success': False, 'message': '缺少或错误的轮次参数'}), 400

        session = get_db_session()
        interview = session.query(Interview).filter(Interview.id == interview_id).first()
        if not interview:
            session.close()
            return jsonify({'success': False, 'message': '面试记录不存在'}), 404

        # 选择对应轮次的文档路径
        doc_path = None
        if round_str == '1':
            doc_path = interview.round1_doc_path
        elif round_str == '2':
            doc_path = interview.round2_doc_path
        else:
            doc_path = interview.round3_doc_path

        if not doc_path:
            session.close()
            return jsonify({'success': False, 'message': '当前轮次暂无文档可供分析'}), 400

        file_path = os.path.join(app.static_folder, doc_path)
        if not os.path.exists(file_path):
            session.close()
            return jsonify({'success': False, 'message': '文档文件不存在，请重新上传'}), 400

        # 获取有效的AI配置（优先级：用户session > 全局配置 > 环境变量）
        ai_config = get_effective_ai_config()
        
        # 提取文档文本
        try:
            doc_text = extract_text(file_path)
        except Exception as e:
            session.close()
            return jsonify({'success': False, 'message': f'文档内容提取失败: {str(e)}'}), 500

        if not doc_text:
            session.close()
            return jsonify({'success': False, 'message': '文档内容为空，无法分析'}), 400

        # 使用有效的AI配置创建提取器
        ai_extractor = create_ai_extractor(ai_config)
        
        if not ai_extractor:
            session.close()
            return jsonify({'success': False, 'message': 'AI功能未启用或未配置API密钥，请在设置中配置AI'}), 400

        # 读取岗位信息（用于结合岗位要求分析）
        position_info_text = ""
        try:
            resume = session.query(Resume).filter(Resume.id == interview.resume_id).first()
            if resume and resume.applied_position:
                position = session.query(Position).filter(Position.position_name == resume.applied_position).first()
                if position:
                    position_info_text = f"""
【岗位信息】
岗位名称：{position.position_name}
工作内容：{position.work_content or '未填写'}
任职资格：{position.job_requirements or '未填写'}
核心需求：{position.core_requirements or '未填写'}
"""
        except Exception as _:
            position_info_text = ""

        # 构建分析提示词
        round_name = {'1': '一面', '2': '二面', '3': '三面'}[round_str]
        prompt = f"""请你作为一名资深HR，根据以下【{round_name}面试录音逐字稿】内容，以及岗位要求，对候选人的表现进行专业分析。

{position_info_text}

【分析要求】：
1. 用 3-5 条要点总结候选人的核心优点（专业能力、沟通表达、思维方式、价值观等）。
2. 用 3-5 条要点指出候选人的主要不足或风险点。
3. 给出一个综合评价结论（适合/可考虑/不太适合），并简要说明原因。
4. 结合上面的岗位信息，特别指出候选人与岗位在关键要求上的高度匹配点和明显不匹配点。
5. 给出5个可用于下一轮面试的重点追问问题，问题要结合本轮表现和岗位要求设计。

【面试逐字稿】：
{doc_text}

请用中文输出，采用以下JSON格式返回（不要包含额外解释文字）：
{{
  "summary": "整体概括（3-5句话）",
  "strengths": ["优点1", "优点2"],
  "weaknesses": ["不足1", "不足2"],
  "conclusion": "综合结论，例如：总体匹配度较高，建议进入下一轮/可以考虑/不太适合等",
  "next_questions": ["下一轮追问问题1", "下一轮追问问题2"]
}}"""

        try:
            response_text = ai_extractor._call_ai_api(prompt)

            # 解析JSON
            import re
            try:
                json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
                if json_match:
                    analysis = json.loads(json_match.group())
                else:
                    analysis = json.loads(response_text)
            except json.JSONDecodeError:
                analysis = {
                    'summary': response_text[:500] if len(response_text) > 500 else response_text,
                    'strengths': [],
                    'weaknesses': [],
                    'conclusion': '',
                    'next_questions': []
                }

            session.close()
            return jsonify({'success': True, 'data': analysis})
        except Exception as e:
            session.close()
            return jsonify({'success': False, 'message': f'AI分析失败: {str(e)}'}), 500

    except Exception as e:
        return jsonify({'success': False, 'message': f'分析请求失败: {str(e)}'}), 500


@app.route('/api/interviews/<int:interview_id>/registration-form', methods=['PUT'])
def update_registration_form(interview_id):
    """更新面试登记表"""
    data = request.json or {}
    session = get_db_session()
    try:
        interview = session.query(Interview).filter(Interview.id == interview_id).first()
        if not interview:
            return jsonify({'success': False, 'message': '面试记录不存在'}), 404
        
        # 更新面试登记表字段
        if 'registration_form_fill_date' in data:
            interview.registration_form_fill_date = data.get('registration_form_fill_date')
        if 'registration_form_contact' in data:
            interview.registration_form_contact = _normalize_field(
                data.get('registration_form_contact'))
        if 'registration_form_email' in data:
            interview.registration_form_email = _normalize_field(
                data.get('registration_form_email'))
        if 'registration_form_birth_date' in data:
            interview.registration_form_birth_date = _normalize_field(
                data.get('registration_form_birth_date'))
        if 'registration_form_ethnicity' in data:
            interview.registration_form_ethnicity = _normalize_field(
                data.get('registration_form_ethnicity'))
        if 'registration_form_marital_status' in data:
            interview.registration_form_marital_status = _normalize_field(
                data.get('registration_form_marital_status'))
        if 'registration_form_has_children' in data:
            interview.registration_form_has_children = _normalize_field(
                data.get('registration_form_has_children'))
        if 'registration_form_origin' in data:
            interview.registration_form_origin = _normalize_field(
                data.get('registration_form_origin'))
        if 'registration_form_id_card' in data:
            interview.registration_form_id_card = _normalize_field(
                data.get('registration_form_id_card'))
        if 'registration_form_first_work_date' in data:
            interview.registration_form_first_work_date = _normalize_field(
                data.get('registration_form_first_work_date'))
        if 'registration_form_education_start_date' in data:
            interview.registration_form_education_start_date = _normalize_field(
                data.get('registration_form_education_start_date'))
        if 'registration_form_education_end_date' in data:
            interview.registration_form_education_end_date = _normalize_field(
                data.get('registration_form_education_end_date'))
        if 'registration_form_institution' in data:
            interview.registration_form_institution = _normalize_field(
                data.get('registration_form_institution'))
        if 'registration_form_major' in data:
            interview.registration_form_major = _normalize_field(
                data.get('registration_form_major'))
        if 'registration_form_degree' in data:
            interview.registration_form_degree = _normalize_field(
                data.get('registration_form_degree'))
        if 'registration_form_full_time' in data:
            interview.registration_form_full_time = _normalize_field(
                data.get('registration_form_full_time'))
        if 'registration_form_recent_work_experience' in data:
            work_exps = data.get('registration_form_recent_work_experience')
            if work_exps is not None:
                interview.registration_form_recent_work_experience = json.dumps(work_exps, ensure_ascii=False)
        if 'registration_form_education_start_date' in data:
            interview.registration_form_education_start_date = _normalize_field(
                data.get('registration_form_education_start_date'))
        if 'registration_form_education_end_date' in data:
            interview.registration_form_education_end_date = _normalize_field(
                data.get('registration_form_education_end_date'))
        if 'registration_form_institution' in data:
            interview.registration_form_institution = _normalize_field(
                data.get('registration_form_institution'))
        if 'registration_form_major' in data:
            interview.registration_form_major = _normalize_field(
                data.get('registration_form_major'))
        if 'registration_form_degree' in data:
            interview.registration_form_degree = _normalize_field(
                data.get('registration_form_degree'))
        if 'registration_form_full_time' in data:
            interview.registration_form_full_time = _normalize_field(
                data.get('registration_form_full_time'))
        if 'registration_form_education' in data:
            interview.registration_form_education = data.get('registration_form_education')
        if 'registration_form_hobbies' in data:
            interview.registration_form_hobbies = data.get('registration_form_hobbies')
        if 'registration_form_current_salary' in data:
            interview.registration_form_current_salary = data.get('registration_form_current_salary')
        if 'registration_form_expected_salary' in data:
            interview.registration_form_expected_salary = data.get('registration_form_expected_salary')
        if 'registration_form_available_date' in data:
            interview.registration_form_available_date = data.get('registration_form_available_date')
        if 'registration_form_address_province' in data:
            interview.registration_form_address_province = data.get('registration_form_address_province')
        if 'registration_form_address_city' in data:
            interview.registration_form_address_city = data.get('registration_form_address_city')
        if 'registration_form_address_district' in data:
            interview.registration_form_address_district = data.get('registration_form_address_district')
        if 'registration_form_address_detail' in data:
            interview.registration_form_address_detail = data.get('registration_form_address_detail')
        if 'registration_form_can_travel' in data:
            interview.registration_form_can_travel = data.get('registration_form_can_travel')
        if 'registration_form_consideration_factors' in data:
            factors = data.get('registration_form_consideration_factors')
            # 确保将列表转换为JSON字符串，即使是空列表也要转换为'[]'
            if factors is not None:
                interview.registration_form_consideration_factors = json.dumps(factors, ensure_ascii=False)
            else:
                interview.registration_form_consideration_factors = ''
        
        # 如果填写日期为空，自动设置为当前日期
        if not interview.registration_form_fill_date:
            interview.registration_form_fill_date = datetime.now().strftime('%Y-%m-%d')
        
        session.commit()
        return jsonify({'success': True, 'data': interview.to_dict()})
    except Exception as e:
        session.rollback()
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'}), 500
    finally:
        session.close()


@app.route('/api/interviews/<int:interview_id>/registration-form-link', methods=['POST'])
def generate_registration_form_link(interview_id):
    """生成面试登记表填写链接"""
    session = get_db_session()
    try:
        interview = session.query(Interview).filter(Interview.id == interview_id).first()
        if not interview:
            return jsonify({'success': False, 'message': '面试记录不存在'}), 404
        
        # 生成唯一token
        token = secrets.token_urlsafe(32)
        interview.registration_form_token = token
        
        session.commit()
        return jsonify({'success': True, 'data': {'token': token}})
    except Exception as e:
        session.rollback()
        return jsonify({'success': False, 'message': f'生成链接失败: {str(e)}'}), 500
    finally:
        session.close()


@app.route('/registration-form', methods=['GET'])
def registration_form_page():
    """面试登记表填写页面（公开页面）"""
    token = request.args.get('token', '').strip()
    if not token:
        return render_template('error.html', message='缺少访问令牌'), 400
    
    session = get_db_session()
    try:
        # 查询对象
        interview = session.query(Interview).filter(Interview.registration_form_token == token).first()
        if not interview:
            return render_template('error.html', message='无效的访问令牌'), 404
        
        # 获取关联的简历信息
        resume = session.query(Resume).filter(Resume.id == interview.resume_id).first()
        
        # 先保存原始值（JSON字符串）
        original_work_exp = interview.registration_form_recent_work_experience
        original_factors = interview.registration_form_consideration_factors
        
        # 将对象从session中分离，避免触发数据库更新
        session.expunge(interview)
        if resume:
            session.expunge(resume)
        
        # 使用 make_transient 确保对象完全脱离 session，不会触发任何数据库操作
        make_transient(interview)
        if resume:
            make_transient(resume)
        
        # 现在可以安全地修改属性，因为对象已完全脱离 session
        # 解析 JSON 展示（对象已完全脱离 session，不会触发数据库更新）
        interview.registration_form_recent_work_experience = interview._parse_json_field(
            original_work_exp, [])
        interview.registration_form_consideration_factors = interview._parse_json_field(
            original_factors, [])
        
        return render_template('registration_form.html', 
                             interview=interview,
                             resume=resume,
                             token=token)
    finally:
        session.close()


def _normalize_field(value):
    if isinstance(value, dict):
        return value.get('value') or value.get('label') or ''
    return value


@app.route('/api/registration-form/submit', methods=['POST'])
def submit_registration_form():
    """提交面试登记表（公开接口）"""
    data = request.json or {}
    token = data.get('token', '').strip()
    
    if not token:
        return jsonify({'success': False, 'message': '缺少访问令牌'}), 400
    
    session = get_db_session()
    try:
        interview = session.query(Interview).filter(Interview.registration_form_token == token).first()
        if not interview:
            return jsonify({'success': False, 'message': '无效的访问令牌'}), 404
        
        # 更新面试登记表字段
        if 'registration_form_contact' in data:
            interview.registration_form_contact = _normalize_field(
                data.get('registration_form_contact'))
        if 'registration_form_email' in data:
            interview.registration_form_email = _normalize_field(
                data.get('registration_form_email'))
        if 'registration_form_birth_date' in data:
            interview.registration_form_birth_date = _normalize_field(
                data.get('registration_form_birth_date'))
        if 'registration_form_ethnicity' in data:
            interview.registration_form_ethnicity = _normalize_field(
                data.get('registration_form_ethnicity'))
        if 'registration_form_marital_status' in data:
            interview.registration_form_marital_status = _normalize_field(
                data.get('registration_form_marital_status'))
        if 'registration_form_has_children' in data:
            interview.registration_form_has_children = _normalize_field(
                data.get('registration_form_has_children'))
        if 'registration_form_origin' in data:
            interview.registration_form_origin = _normalize_field(
                data.get('registration_form_origin'))
        if 'registration_form_id_card' in data:
            interview.registration_form_id_card = _normalize_field(
                data.get('registration_form_id_card'))
        if 'registration_form_recent_work_experience' in data:
            work_exps = data.get('registration_form_recent_work_experience')
            if work_exps is not None:
                interview.registration_form_recent_work_experience = json.dumps(work_exps, ensure_ascii=False)
        if 'registration_form_education' in data:
            interview.registration_form_education = data.get('registration_form_education')
        if 'registration_form_hobbies' in data:
            interview.registration_form_hobbies = data.get('registration_form_hobbies')
        if 'registration_form_current_salary' in data:
            interview.registration_form_current_salary = data.get('registration_form_current_salary')
        if 'registration_form_expected_salary' in data:
            interview.registration_form_expected_salary = data.get('registration_form_expected_salary')
        if 'registration_form_available_date' in data:
            interview.registration_form_available_date = data.get('registration_form_available_date')
        if 'registration_form_address_province' in data:
            interview.registration_form_address_province = data.get('registration_form_address_province')
        if 'registration_form_address_city' in data:
            interview.registration_form_address_city = data.get('registration_form_address_city')
        if 'registration_form_address_district' in data:
            interview.registration_form_address_district = data.get('registration_form_address_district')
        if 'registration_form_address_detail' in data:
            interview.registration_form_address_detail = data.get('registration_form_address_detail')
        if 'registration_form_can_travel' in data:
            interview.registration_form_can_travel = data.get('registration_form_can_travel')
        if 'registration_form_consideration_factors' in data:
            factors = data.get('registration_form_consideration_factors')
            # 确保将列表转换为JSON字符串，即使是空列表也要转换为'[]'
            if factors is not None:
                interview.registration_form_consideration_factors = json.dumps(factors, ensure_ascii=False)
            else:
                interview.registration_form_consideration_factors = ''
        
        # 自动设置填写日期
        interview.registration_form_fill_date = datetime.now().strftime('%Y-%m-%d')
        
        session.commit()
        return jsonify({'success': True, 'message': '信息提交成功'})
    except Exception as e:
        session.rollback()
        return jsonify({'success': False, 'message': f'提交失败: {str(e)}'}), 500
    finally:
        session.close()


@app.route('/api/interviews/<int:interview_id>/analysis-pdf', methods=['POST'])
def export_interview_round_analysis_pdf(interview_id):
    """导出单轮面试AI分析结果为PDF"""
    try:
        data = request.json or {}
        round_str = str(data.get('round') or '')
        if round_str not in ('1', '2', '3'):
            return jsonify({'success': False, 'message': '缺少或错误的轮次参数'}), 400

        session = get_db_session()
        try:
            interview = session.query(Interview).filter(Interview.id == interview_id).first()
        finally:
            session.close()

        if not interview:
            return jsonify({'success': False, 'message': '面试记录不存在'}), 404

        # 选择对应轮的AI分析文本（优先使用数据库中的内容，如为空则尝试使用前端传来的analysis_text）
        client_text = data.get('analysis_text') or ''
        if round_str == '1':
            analysis_text = interview.round1_ai_result or client_text
            round_name = '一面'
        elif round_str == '2':
            analysis_text = interview.round2_ai_result or client_text
            round_name = '二面'
        else:
            analysis_text = interview.round3_ai_result or client_text
            round_name = '三面'

        if not analysis_text:
            return jsonify({'success': False, 'message': '当前轮次暂无AI分析结果，请先执行AI分析'}), 400

        file_path = export_interview_round_analysis_to_pdf(interview, round_name, analysis_text)
        download_name = f'{round_name}面试反馈报告_{interview.name or interview_id}.pdf'
        return send_file(file_path, as_attachment=True, download_name=download_name)
    except Exception as e:
        return jsonify({'success': False, 'message': f'导出AI分析报告失败: {str(e)}'}), 500

@app.route('/api/ai/config', methods=['GET'])
def get_ai_config():
    """获取AI配置（不返回密钥）"""
    ai_enabled = app.config.get('AI_ENABLED', True)
    ai_api_key = app.config.get('AI_API_KEY', '')
    # 检查AI是否真正可用（启用且有API密钥）
    ai_available = ai_enabled and bool(ai_api_key)
    
    return jsonify({
        'success': True,
        'data': {
            'ai_enabled': ai_enabled,
            'ai_available': ai_available,  # 新增：AI是否真正可用
            'ai_model': app.config.get('AI_MODEL', 'gpt-3.5-turbo'),
            'ai_api_base': app.config.get('AI_API_BASE', ''),
            'ai_models': app.config.get('AI_MODELS', [])
        }
    })

@app.route('/api/ai/config', methods=['POST'])
def save_ai_config():
    """保存AI配置"""
    try:
        data = request.json
        ai_enabled = data.get('ai_enabled', True)
        ai_model = data.get('ai_model', 'gpt-3.5-turbo')
        ai_api_key = data.get('ai_api_key', '')
        ai_api_base = data.get('ai_api_base', '')
        
        # 更新配置（注意：这里只是临时更新，重启后会恢复）
        # 实际生产环境应该保存到配置文件或数据库
        app.config['AI_ENABLED'] = ai_enabled
        app.config['AI_MODEL'] = ai_model
        if ai_api_key:
            app.config['AI_API_KEY'] = ai_api_key
        if ai_api_base:
            app.config['AI_API_BASE'] = ai_api_base
        
        return jsonify({
            'success': True,
            'message': 'AI配置已保存（当前会话有效）'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'保存配置失败: {str(e)}'
        }), 400

@app.route('/api/ai/test', methods=['POST'])
def test_ai_connection():
    """测试AI连接"""
    try:
        data = request.json
        api_key = data.get('api_key', '')
        api_base = data.get('api_base', '')
        model = data.get('model', 'gpt-3.5-turbo')
        
        if not api_key:
            return jsonify({
                'success': False,
                'message': '请提供API密钥'
            }), 400
        
        # 创建临时AI提取器进行测试
        ai_extractor = AIExtractor(
            api_key=api_key,
            api_base=api_base if api_base else None,
            model=model
        )
        
        # 使用简单的测试文本
        test_text = "姓名：张三\n性别：男\n手机：13800138000"
        result = ai_extractor.extract_with_ai(test_text)
        
        if result:
            return jsonify({
                'success': True,
                'message': 'AI连接测试成功',
                'data': result
            })
        else:
            return jsonify({
                'success': False,
                'message': 'AI连接测试失败，请检查API密钥和网络连接'
            }), 400
            
    except Exception as e:
            return jsonify({
                'success': False,
                'message': f'测试失败: {str(e)}'
            }), 400


def _collect_registration_data(interview):
    return {
        'name': interview.name or '',
        'gender': interview.registration_form_ethnicity or '',
        'birth_date': interview.registration_form_birth_date or '',
        'marital_status': interview.registration_form_marital_status or '',
        'has_children': interview.registration_form_has_children or '',
        'origin': interview.registration_form_origin or '',
        'id_card': interview.registration_form_id_card or '',
        'first_work_date': interview.registration_form_first_work_date or '',
        'education_start_date': interview.registration_form_education_start_date or '',
        'education_end_date': interview.registration_form_education_end_date or '',
        'institution': interview.registration_form_institution or '',
        'major': interview.registration_form_major or '',
        'degree': interview.registration_form_degree or '',
        'full_time': interview.registration_form_full_time or '',
        'education': interview.registration_form_education or '',
        'hobbies': interview.registration_form_hobbies or '',
        'current_salary': interview.registration_form_current_salary or '',
        'expected_salary': interview.registration_form_expected_salary or '',
        'available_date': interview.registration_form_available_date or '',
        'address': f"{interview.registration_form_address_province or ''} {interview.registration_form_address_city or ''} {interview.registration_form_address_district or ''}",
        'address_detail': interview.registration_form_address_detail or '',
        'can_travel': interview.registration_form_can_travel or '',
        'contact': interview.registration_form_contact or '',
        'email': interview.registration_form_email or '',
        'applied_position': interview.applied_position or '',
        'fill_date': interview.registration_form_fill_date or '',
        'work_experience': interview._parse_json_field(
            interview.registration_form_recent_work_experience, [])[:2],
        'factors': interview._parse_json_field(
            interview.registration_form_consideration_factors, [])
    }




def export_registration_form_to_excel(interview):
    try:
        data = _collect_registration_data(interview)
        wb = Workbook()
        ws = wb.active
        ws.title = '面试登记表'
        ws.sheet_properties.tabColor = "1072BA"

        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        bold = Font(bold=True, size=11)
        title_font = Font(bold=True, size=16)
        section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        for col in ['A','B','C','D','E','F']:
            ws.column_dimensions[col].width = 13

        def stylize_cell(cell, font=None, align=None, fill=None):
            _stylize_cell(cell, border, font, align, fill)

        ws.row_dimensions[1].height = 50
        ws.merge_cells('A1:F1')
        ws['A1'].value = '应聘人员面试登记表'
        stylize_cell(ws['A1'], font=title_font, align=center)

        ws.row_dimensions[2].height = 25
        ws.merge_cells('A2:B2')
        ws.merge_cells('C2:D2')
        ws['A2'].value = f"应聘岗位：{data['applied_position']}"
        ws['C2'].value = f"填表日期：{data['fill_date']}"
        for cell in ['A2','C2']:
            stylize_cell(ws[cell], font=bold, align=center)

        def section_row(row, title):
            ws.row_dimensions[row].height = 25
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1)
            cell.value = title
            stylize_cell(cell, font=bold, align=center, fill=section_fill)

        section_row(3, '个人基本信息')
        _fill_row(ws, 4, ['姓名', data['name'], '性别', data['gender'], '出生年月', data['birth_date']], border, center)
        _fill_row(ws, 5, ['民族', data['origin'], '婚姻状况', data['marital_status'], '有无子女', data['has_children']], border, center)
        _fill_row(ws, 6, ['学历', data['degree'], '联系电话', data['contact'], '电子邮箱', data['email']], border, center)
        _fill_row(ws, 7, ['籍贯', data['origin'], '身份证号', data['id_card'], '', ''], border, center)

        section_row(8, '教育经历')
        _fill_row(ws, 9, ['开始时间', '结束时间', '毕业院校', '专业', '学历', '是否为全日制统招'], border, center)
        _fill_row(ws, 10, [
            data['education_start_date'],
            data['education_end_date'],
            data['institution'],
            data['major'],
            data['degree'],
            data['full_time']
        ], border, center)

        section_row(12, '工作经历（从最近开始）')
        _fill_row(ws, 13, ['开始时间', '结束时间', '单位名称', '', '职务', '离职原因'], border, center)
        ws.merge_cells('C13:D13')
        def job_list_row(row, entry):
            _fill_row(ws, row, [
                entry.get('start_year', ''),
                entry.get('end_year', ''),
                entry.get('company', ''),
                '',
                entry.get('position', ''),
                entry.get('departure_reason', '')
            ], border, center)
            ws.merge_cells(f'C{row}:D{row}')
        job_list_row(14, data['work_experience'][0] if data['work_experience'] else {})
        job_list_row(15, data['work_experience'][1] if len(data['work_experience']) > 1 else {})
        _fill_row(ws, 16, ['', '', '', '', '', ''], border, center)

        section_row(17, '考虑新公司主要原因')
        factor_texts = [
            '公司前景',
            '团队氛围',
            '薪酬福利',
            '人际关系',
            '文化价值观',
            '晋升机会',
            '领导风格'
        ]
        factors = data['factors'] or []
        factor_row = []
        for idx, defaultText in enumerate(factor_texts):
            value = factors[idx] if idx < len(factors) and factors[idx] else defaultText
            factor_row.append(f"{idx + 1}、{value}")
        _fill_row(ws, 18, factor_row[:6], border, center)
        _fill_row(ws, 19, [factor_row[6], '', '', '', '', ''], border, center)

        section_row(21, '个人爱好及专长')
        ws.row_dimensions[22].height = 25
        ws.merge_cells('A22:F22')
        ws['A22'].value = data['hobbies']
        stylize_cell(ws['A22'], align=left)

        _fill_row(ws, 23, ['原月薪', data['current_salary'], '期望月薪', data['expected_salary'], '最快到岗时间', data['available_date']], border, center)
        ws['A24'].value = '现住址'
        ws.merge_cells('B24:F24')
        ws['B24'].value = f"{data['address']} {data['address_detail']}"
        stylize_cell(ws['B24'], align=center)
        ws.row_dimensions[24].height = 25
        for r in range(25, 30):
            ws.row_dimensions[r].height = 25
        ws.merge_cells('A25:F29')
        ws['A25'].value = (
            "声明人：\n"
            "\n"
            "本人保证以上内容填写真实无误，同意接受公司的调查，若有不实之处，本人愿意承担一切后果。\n"
            "\n"
            "声明人签字\n"
            "\n"
            "\n"
            "日期"
        )
        stylize_cell(ws['A25'], align=center)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        print("Excel 导出失败：", file=sys.stderr)
        traceback.print_exc()
        raise



def export_registration_form_to_pdf(interview):
    data = _collect_registration_data(interview)
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.setFont('STSong-Light', 16)
    c.drawCentredString(width / 2, height - 40, '应聘人员面试登记表')

    y = height - 80
    def draw_section(title, rows):
        nonlocal y
        c.setFont('STSong-Light', 12)
        c.drawString(40, y, title)
        y -= 20
        table = Table(rows, colWidths=[(width-80)/len(rows[0])] * len(rows[0]))
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, -1), 'STSong-Light'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        table.wrapOn(c, width - 80, y)
        table.drawOn(c, 40, y - 10 - (20 * len(rows)))
        y -= 30 + 20 * len(rows)

    draw_section('个人基本信息', [
        ['姓名', data['name'], '性别', data['gender'], '联系方式', data['contact']],
        ['出生日期', data['birth_date'], '民族', data['origin'], '婚姻状况', data['marital_status']],
        ['籍贯', data['origin'], '身份证号', data['id_card'], '邮箱', data['email']],
    ])
    draw_section('教育/自我信息', [
        ['最高学历', data['education'], '学位', data['degree'], '统招', data['full_time']],
        ['毕业院校', data['institution'], '专业', data['major'], '起止时间', f"{data['education_start_date']} - {data['education_end_date']}"],
        ['个人爱好及特长', data['hobbies'], '原月薪', data['current_salary'], '期望月薪', data['expected_salary']],
        ['最快到岗时间', data['available_date'], '能否出差', data['can_travel'], '', '']
    ])
    draw_section('现住址', [
        ['省/市/区', data['address'], '详细地址', data['address_detail'], '', '']
    ])
    exp_rows = [['公司名称', '岗位', '开始时间', '结束时间']]
    for exp in data['work_experience']:
        exp_rows.append([
            exp.get('company', ''),
            exp.get('position', ''),
            exp.get('start_year') or '',
            exp.get('end_year') or ''
        ])
    if len(exp_rows) == 1:
        exp_rows.append(['暂无工作经历', '', '', ''])
    draw_section('近两份工作经历', exp_rows)
    factor_rows = [['排序结果']] + [[f'{idx + 1}. {factor}'] for idx, factor in enumerate(data['factors'] or ['未填写'])]
    draw_section('考虑新公司的主要因素', factor_rows)

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer


@app.route('/api/interviews/<int:interview_id>/registration-form/export', methods=['GET'])
def export_registration_form(interview_id):
    fmt = request.args.get('format', 'excel').lower()
    session = get_db_session()
    try:
        interview = session.query(Interview).filter(Interview.id == interview_id).first()
        if not interview:
            return jsonify({'success': False, 'message': '面试记录不存在'}), 404
        if fmt == 'pdf':
            pdf_file = export_registration_form_to_pdf(interview)
            return send_file(
                pdf_file,
                as_attachment=True,
                download_name=f'面试登记表_{interview.name or interview_id}.pdf',
                mimetype='application/pdf'
            )
        excel_file = export_registration_form_to_excel(interview)
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f'面试登记表_{interview.name or interview_id}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500
    finally:
        session.close()

# 岗位目录API
@app.route('/api/positions', methods=['GET'])
def get_positions():
    """获取岗位列表"""
    try:
        session = get_db_session()
        positions = session.query(Position).order_by(Position.update_time.desc()).all()
        session.close()
        
        return jsonify({
            'success': True,
            'data': [pos.to_dict() for pos in positions]
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取岗位列表失败: {str(e)}'
        }), 500

@app.route('/api/positions', methods=['POST'])
def create_position():
    """创建岗位"""
    try:
        data = request.json
        position_name = data.get('position_name', '').strip()
        
        if not position_name:
            return jsonify({
                'success': False,
                'message': '岗位名称不能为空'
            }), 400
        
        session = get_db_session()
        current_user = get_current_user()
        username = current_user.username if current_user else 'system'
        
        position = Position(
            position_name=position_name,
            work_content=data.get('work_content', '').strip() or None,
            job_requirements=data.get('job_requirements', '').strip() or None,
            core_requirements=data.get('core_requirements', '').strip() or None,
            created_by=username,
            updated_by=username
        )
        session.add(position)
        session.commit()
        position_id = position.id
        session.close()
        
        return jsonify({
            'success': True,
            'message': '岗位创建成功',
            'data': position.to_dict()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'创建岗位失败: {str(e)}'
        }), 500

@app.route('/api/positions/<int:position_id>', methods=['GET'])
def get_position(position_id):
    """获取单个岗位详情"""
    try:
        session = get_db_session()
        position = session.query(Position).filter(Position.id == position_id).first()
        session.close()
        
        if not position:
            return jsonify({
                'success': False,
                'message': '岗位不存在'
            }), 404
        
        return jsonify({
            'success': True,
            'data': position.to_dict()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'获取岗位详情失败: {str(e)}'
        }), 500

@app.route('/api/positions/<int:position_id>', methods=['PUT'])
def update_position(position_id):
    """更新岗位"""
    try:
        data = request.json
        position_name = data.get('position_name', '').strip()
        
        if not position_name:
            return jsonify({
                'success': False,
                'message': '岗位名称不能为空'
            }), 400
        
        session = get_db_session()
        position = session.query(Position).filter(Position.id == position_id).first()
        
        if not position:
            session.close()
            return jsonify({
                'success': False,
                'message': '岗位不存在'
            }), 404
        
        current_user = get_current_user()
        username = current_user.username if current_user else 'system'
        
        position.position_name = position_name
        position.work_content = data.get('work_content', '').strip() or None
        position.job_requirements = data.get('job_requirements', '').strip() or None
        position.core_requirements = data.get('core_requirements', '').strip() or None
        position.update_time = datetime.now()
        position.updated_by = username
        
        session.commit()
        session.close()
        
        return jsonify({
            'success': True,
            'message': '岗位更新成功',
            'data': position.to_dict()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'更新岗位失败: {str(e)}'
        }), 500

@app.route('/api/positions/<int:position_id>', methods=['DELETE'])
@admin_required
def delete_position(position_id):
    """删除岗位（仅管理员）"""
    try:
        session = get_db_session()
        position = session.query(Position).filter(Position.id == position_id).first()
        
        if not position:
            session.close()
            return jsonify({
                'success': False,
                'message': '岗位不存在'
            }), 404
        
        session.delete(position)
        session.commit()
        session.close()
        
        return jsonify({
            'success': True,
            'message': '岗位删除成功'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'删除岗位失败: {str(e)}'
        }), 500

@app.route('/api/sync/check', methods=['GET'])
@login_required
def check_sync():
    """检查数据更新（用于实时同步）"""
    try:
        last_sync_time = request.args.get('last_sync', '')
        current_user = get_current_user()
        if not current_user:
            return jsonify({'success': False, 'message': '未登录'}), 401
        
        db = get_db_session()
        try:
            updates = {
                'resumes': False,
                'interviews': False,
                'positions': False
            }
            
            if last_sync_time:
                try:
                    last_sync = datetime.fromisoformat(last_sync_time.replace('Z', '+00:00'))
                except:
                    last_sync = None
                
                if last_sync:
                    # 检查简历更新
                    resume_count = db.query(Resume).filter(
                        Resume.updated_at > last_sync
                    ).count()
                    if resume_count > 0:
                        updates['resumes'] = True
                    
                    # 检查面试流程更新
                    interview_count = db.query(Interview).filter(
                        Interview.updated_at > last_sync
                    ).count()
                    if interview_count > 0:
                        updates['interviews'] = True
                    
                    # 检查岗位更新
                    position_count = db.query(Position).filter(
                        Position.updated_at > last_sync
                    ).count()
                    if position_count > 0:
                        updates['positions'] = True
            
            return jsonify({
                'success': True,
                'updates': updates,
                'current_time': datetime.now().isoformat()
            })
        finally:
            db.close()
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/resumes/<int:resume_id>/match-analysis', methods=['POST'])
def analyze_resume_match(resume_id):
    """分析简历与岗位的匹配度"""
    try:
        data = request.json
        applied_position = data.get('applied_position', '').strip()
        
        if not applied_position:
            return jsonify({
                'success': False,
                'message': '请先选择应聘岗位'
            }), 400
        
        # 获取简历信息
        session = get_db_session()
        resume = session.query(Resume).filter(Resume.id == resume_id).first()
        if not resume:
            session.close()
            return jsonify({
                'success': False,
                'message': '简历不存在'
            }), 404
        
        # 获取岗位信息
        position = session.query(Position).filter(Position.position_name == applied_position).first()
        if not position:
            session.close()
            return jsonify({
                'success': False,
                'message': '岗位不存在，请先在岗位目录中添加该岗位'
            }), 404
        
        session.close()
        
        # 获取有效的AI配置（优先级：用户session > 全局配置 > 环境变量）
        ai_config = get_effective_ai_config()
        
        # 创建AI提取器
        ai_extractor = create_ai_extractor(ai_config)
        
        if not ai_extractor:
            return jsonify({
                'success': False,
                'message': 'AI功能未启用或未配置API密钥，请在设置中配置AI'
            }), 400
        
        # 构建分析提示
        resume_info = f"""
姓名：{resume.name or '未知'}
性别：{resume.gender or '未知'}
年龄：{resume.age or '未知'}
学历：{resume.highest_education or '未知'}
毕业学校：{resume.school or '未知'}
专业：{resume.major or '未知'}
工龄：{resume.earliest_work_year and (datetime.now().year - resume.earliest_work_year) or '未知'}年
工作经历：{json.dumps(resume.work_experience or [], ensure_ascii=False, indent=2)}
"""
        
        position_info = f"""
岗位名称：{position.position_name}
工作内容：{position.work_content or '未填写'}
任职资格：{position.job_requirements or '未填写'}
核心需求：{position.core_requirements or '未填写'}
"""
        
        prompt = f"""请分析以下简历与岗位的匹配度，并给出详细的分析报告。

【简历信息】
{resume_info}

【岗位要求】
{position_info}

请从以下维度进行分析：
1. 教育背景匹配度（学历、学校、专业）
2. 工作经验匹配度（工作年限、工作内容、岗位相关性）
3. 技能匹配度（根据工作经历推断的技能）
4. 综合匹配度评分（0-100分）

请以JSON格式返回分析结果，格式如下：
{{
    "match_score": 85,
    "match_level": "高度匹配",
    "detailed_analysis": "详细的分析说明...",
    "strengths": ["优势1", "优势2"],
    "weaknesses": ["不足1", "不足2"],
    "suggestions": ["【考核重点】技术能力 - 【面试问题】请详细说明您在XX项目中的技术实现方案和遇到的挑战", "【考核重点】沟通协作 - 【面试问题】请描述一次您与跨部门团队协作解决复杂问题的经历"]
}}

其中：
- match_score: 匹配度分数（0-100）
- match_level: 匹配等级（高度匹配/中等匹配/低度匹配）
- detailed_analysis: 详细分析说明（200-500字）
- strengths: 优势匹配点列表
- weaknesses: 不足匹配点列表
- suggestions: 这是给面试官使用的面试重点考核项及对应面试问题，不是给候选人的建议！

【suggestions字段的详细要求】：
1. 这是给面试官的建议，用于指导面试官在面试中重点考核哪些方面，以及应该问什么问题
2. 绝对不要生成给候选人的改进建议（如"建议候选人如何提升"、"候选人应该做什么"等）
3. 必须严格按照以下格式生成，每个suggestion必须是：【考核重点】考核项名称 - 【面试问题】具体的面试问题
4. 根据简历与岗位的匹配情况，识别3-5个需要重点考核的维度，例如：
   - 如果简历缺乏相关经验，考核重点可以是"行业经验"或"学习能力"
   - 如果简历有相关经验但不够深入，考核重点可以是"项目深度"或"技术能力"
   - 如果岗位需要沟通能力，考核重点可以是"沟通协作"或"团队合作"
5. 为每个考核重点设计1-2个针对性的面试问题，问题要能帮助面试官深入了解候选人在该维度的真实能力
6. 面试问题应该以"请"、"请描述"、"请说明"等开头，直接面向候选人提问

【格式示例】：
正确格式：
- "【考核重点】技术能力 - 【面试问题】请详细说明您在XX项目中的技术实现方案和遇到的挑战"
- "【考核重点】沟通协作 - 【面试问题】请描述一次您与跨部门团队协作解决复杂问题的经历"

错误格式（禁止使用）：
- "建议候选人提升技术能力"（这是给候选人的建议，不是给面试官的）
- "候选人应该加强沟通能力"（这是给候选人的建议，不是给面试官的）
- "技术能力：请说明..."（缺少【考核重点】和【面试问题】标记）

请只返回JSON格式，不要包含其他文字说明。"""
        
        try:
            response_text = ai_extractor._call_ai_api(prompt)
            
            # 尝试解析JSON响应
            try:
                # 提取JSON部分
                import re
                json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
                if json_match:
                    analysis_result = json.loads(json_match.group())
                else:
                    # 如果找不到JSON，尝试直接解析
                    analysis_result = json.loads(response_text)
            except json.JSONDecodeError:
                # 如果解析失败，返回一个默认结构
                analysis_result = {
                    'match_score': 50,
                    'match_level': '中等匹配',
                    'detailed_analysis': response_text[:500] if len(response_text) > 500 else response_text,
                    'strengths': [],
                    'weaknesses': [],
                    'suggestions': []
                }
            
            # 验证和清理suggestions字段，确保格式正确
            if 'suggestions' in analysis_result and isinstance(analysis_result['suggestions'], list):
                import re
                cleaned_suggestions = []
                for suggestion in analysis_result['suggestions']:
                    if not isinstance(suggestion, str):
                        continue
                    # 检查是否符合格式：【考核重点】xxx - 【面试问题】xxx
                    if re.match(r'【考核重点】.*?\s*[-—–]\s*【面试问题】.*', suggestion):
                        cleaned_suggestions.append(suggestion)
                    # 如果不符合格式，尝试修复或跳过
                    # 过滤掉明显是给候选人的建议（包含"建议"、"应该"等关键词）
                    elif any(keyword in suggestion for keyword in ['建议候选人', '候选人应该', '建议您', '您应该', '可以尝试']):
                        # 跳过给候选人的建议
                        continue
                analysis_result['suggestions'] = cleaned_suggestions
            
            # 对匹配得分进行"温和放宽"，避免评分过于严苛
            # 原始得分区间 0-100，转换为约 50-100 的区间，更贴近日常使用场景
            try:
                raw_score = analysis_result.get('match_score')
                if raw_score is None:
                    raw_score = 60
                raw_score = float(raw_score)
                # 简单线性放宽：new = raw * 0.7 + 30，限制在 [50, 100]
                new_score = int(max(50, min(100, raw_score * 0.7 + 30)))
                analysis_result['match_score'] = new_score

                # 根据调整后的得分重新划分匹配等级
                # ≥80：高匹配度；≥60：中匹配度；<60：低匹配度
                if new_score >= 80:
                    level = '高度匹配'
                elif new_score >= 60:
                    level = '中等匹配'
                else:
                    level = '低度匹配'
                analysis_result['match_level'] = level
            except Exception:
                # 若转换失败，则保持原始结果
                pass

            # 保存匹配结果到简历记录
            try:
                session_save = get_db_session()
                resume_save = session_save.query(Resume).filter(Resume.id == resume_id).first()
                if resume_save:
                    current_user = get_current_user()
                    username = current_user.username if current_user else 'system'
                    resume_save.match_score = analysis_result.get('match_score')
                    resume_save.match_level = analysis_result.get('match_level')
                    resume_save.match_position = applied_position
                    resume_save.updated_by = username
                    session_save.commit()
                session_save.close()
            except Exception as save_err:
                # 不影响主流程，仅打印日志
                print(f"保存匹配结果到简历记录失败: {save_err}")
            
            # 同步匹配结果到面试流程（如有对应的面试记录）
            try:
                session_sync = get_db_session()
                current_user = get_current_user()
                username = current_user.username if current_user else 'system'
                interviews = session_sync.query(Interview).filter(Interview.resume_id == resume_id).all()
                if interviews:
                    for it in interviews:
                        # 如果面试流程的岗位与分析的岗位一致，则更新匹配度
                        if it.applied_position == applied_position:
                            it.match_score = analysis_result.get('match_score')
                            it.match_level = analysis_result.get('match_level')
                            it.analyzed_by = username  # 记录分析者
                            it.updated_by = username
                    session_sync.commit()
                session_sync.close()
            except Exception as sync_err:
                # 不影响主流程，仅打印日志
                print(f"同步匹配结果到面试流程失败: {sync_err}")
            
            # 添加分析者信息
            current_user = get_current_user()
            username = current_user.username if current_user else 'system'
            analysis_result['analyzed_by'] = username
            
            return jsonify({
                'success': True,
                'data': analysis_result
            })
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'AI分析失败: {str(e)}'
            }), 500
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'分析失败: {str(e)}'
        }), 500

if __name__ == '__main__':
    import socket
    import sys
    
    # 支持生产环境部署（Railway、Render 等）
    # 从环境变量读取端口和主机，如果没有则使用默认值
    port = int(os.environ.get('PORT', 5000))
    # 本地开发默认使用 127.0.0.1，生产环境（Railway）自动使用 0.0.0.0
    # 如果设置了 PORT 环境变量，说明是生产环境，使用 0.0.0.0
    is_production = 'PORT' in os.environ
    host = os.environ.get('HOST', '0.0.0.0' if is_production else '127.0.0.1')
    debug = os.environ.get('DEBUG', 'False' if is_production else 'True').lower() == 'true'
    
    # 只在本地开发环境检查端口占用（生产环境由平台管理）
    if not is_production:
        def is_port_in_use(port):
            """检查端口是否被占用"""
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(1)
                try:
                    # 尝试绑定端口，如果失败说明端口被占用
                    s.bind(('127.0.0.1', port))
                    return False
                except OSError:
                    return True
        
        if is_port_in_use(port):
            print(f'错误: 端口 {port} 已被占用！')
            print('请关闭占用该端口的程序，或修改 app.py 中的端口号。')
            print()
            print('提示：可以通过以下命令查看占用端口的进程：')
            print(f'  netstat -ano | findstr :{port}')
            sys.exit(1)
    
    print('=' * 50)
    print('智能简历数据库系统')
    print('=' * 50)
    if host == '0.0.0.0':
        print(f'服务器地址: http://127.0.0.1:{port}')
        print(f'局域网地址: http://0.0.0.0:{port}')
    else:
        print(f'服务器地址: http://{host}:{port}')
    print(f'调试模式: {debug}')
    print('=' * 50)
    print('按 Ctrl+C 停止服务器')
    print('=' * 50)
    print()
    
    try:
        app.run(debug=debug, host=host, port=port, use_reloader=False)
    except Exception as e:
        print(f'启动失败: {e}')
        import traceback
        traceback.print_exc()
        sys.exit(1)

